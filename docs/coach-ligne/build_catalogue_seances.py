#!/usr/bin/env python3
"""Catalogue coach : 100 séances × 13 familles.

Gold Arthur. Structure : échauffement → bloc de séance → retour au calme.
Nager : pas de phase. Tri / eau libre : 2 tests (T50+T100, T400) puis 4 deload,
puis 94 construction. Pas de doublon. Totaux ×100. Trous {éducatif} et {matériel}.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).with_name("catalogue-seances-13-familles.xlsx")

NAVY, BLUE = "0A162C", "006BFD"
LINE = "D5DEE8"
thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_soft = PatternFill("solid", fgColor="E8F1FF")
fill_alt = PatternFill("solid", fgColor="F7FAFC")
ROLE_FILL = {
    "séance technique": PatternFill("solid", fgColor="D6E4FF"),
    "séance vitesse": PatternFill("solid", fgColor="FFE0B2"),
    "séance endurance": PatternFill("solid", fgColor="C8E6C9"),
    "séance jambes": PatternFill("solid", fgColor="E1BEE7"),
    "séance récupération": PatternFill("solid", fgColor="ECEFF1"),
}
PHASE_FILL = {
    "construction": PatternFill("solid", fgColor="C8E6C9"),
    "test": PatternFill("solid", fgColor="FFE0B2"),
    "deload": PatternFill("solid", fgColor="ECEFF1"),
}
font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_h = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11, color=NAVY)
font_small = Font(name="Calibri", size=10, color=NAVY)
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS_NAGER = [
    "n°",
    "bande",
    "total_m",
    "échauffement",
    "bloc de séance",
    "retour au calme",
    "contrôle_somme",
]
HEADERS_EPREUVE = [
    "n°",
    "phase",
    "bande",
    "total_m",
    "échauffement",
    "bloc de séance",
    "retour au calme",
    "contrôle_somme",
]


def nx(n: int, dist: int, text: str, rest: int | None = None) -> str:
    core = f"{n} × {dist} m {text}" if n > 1 else f"{dist} m {text}"
    if rest is None or n <= 1:
        return core
    return f"{core}, repos {rest} s"


def meters_of(text: str) -> int:
    """Ignore les mètres entre parenthèses. Ne pas compter « 1 min »."""
    cleaned = re.sub(r"\([^)]*\)", " ", text or "")
    total = 0
    for m in re.finditer(r"(\d+)\s*[×x]\s*(\d+)\s*m(?![a-zA-Z])", cleaned, re.I):
        total += int(m.group(1)) * int(m.group(2))
    stripped = re.sub(r"\d+\s*[×x]\s*\d+\s*m(?![a-zA-Z])", " ", cleaned, flags=re.I)
    for m in re.finditer(r"(\d+)\s*m(?![a-zA-Z])", stripped, re.I):
        total += int(m.group(1))
    return total


def join(lines: list[str]) -> str:
    return "\n".join(x for x in lines if x)


def pad(lines: list[str], need: int, text: str, rest: int = 20) -> list[str]:
    lines = [ln for ln in lines if ln]
    while lines and meters_of(join(lines)) > need:
        lines.pop()
    while lines and meters_of(join(lines)) % 50:
        lines.pop()
    have = meters_of(join(lines)) if lines else 0
    extra = need - have
    if extra < 0 or extra % 50:
        raise ValueError(f"pad {have} → {need} extra={extra}")
    if extra == 0:
        return lines
    n100, rem = divmod(extra, 100)
    out = list(lines)
    variant = getattr(pad, "variant", 0)
    if n100 >= 2:
        a = 1 + (variant % n100)
        alt = "crawl souple" if "nage libre" in text else "nage libre souple"
        if a >= n100:
            out.append(nx(n100, 100, text, rest))
        else:
            out.append(nx(a, 100, text, rest))
            out.append(nx(n100 - a, 100, alt, rest))
    elif n100:
        out.append(nx(n100, 100, text, rest))
    if rem == 50:
        out.append(f"50 m {text}")
    return out


BLOCK_DEBUTANT = [
    "séance endurance", "séance technique", "séance endurance", "séance jambes", "séance récupération",
    "séance endurance", "séance technique", "séance endurance", "séance technique", "séance jambes",
]
BLOCK_NAGER = [
    "séance endurance", "séance technique", "séance endurance", "séance vitesse", "séance récupération",
    "séance endurance", "séance jambes", "séance technique", "séance endurance", "séance endurance",
]
BLOCK_SHORT = [
    "séance endurance", "séance vitesse", "séance technique", "séance endurance", "séance jambes",
    "séance endurance", "séance vitesse", "séance technique", "séance récupération", "séance endurance",
]
BLOCK_LONG = [
    "séance endurance", "séance endurance", "séance technique", "séance jambes", "séance récupération",
    "séance endurance", "séance endurance", "séance technique", "séance vitesse", "séance endurance",
]


def playlist(kind: str) -> list[str]:
    block = {
        "debutant": BLOCK_DEBUTANT, "nager": BLOCK_NAGER, "short": BLOCK_SHORT,
        "long": BLOCK_LONG, "ow_short": BLOCK_NAGER, "ow_long": BLOCK_LONG, "four_n": BLOCK_NAGER,
    }[kind]
    return (block * 10)[:100]


def is_event_family(fam: dict) -> bool:
    return fam["flavor"] != "nager"


def spec_for(fam: dict, i: int) -> tuple[str | None, str]:
    """phase affichée (None si Nager) + rôle interne pour le contenu.

    Épreuve : n°1-2 test, n°3-6 deload, n°7-100 construction.
    """
    if not is_event_family(fam):
        return None, playlist(fam["play"])[i]
    if i == 0:
        return "test", "test_50_100"
    if i == 1:
        return "test", "test_400"
    if i <= 5:
        return "deload", f"deload_{i - 1}"
    role = ["séance endurance", "séance technique", "séance endurance", "séance jambes"][(i - 6) % 4]
    return "construction", role


FAMILIES = [
    {"sheet": "01 Nager deb crawl", "title": "Nager · débutant · crawl (+ 25 m dos/brasse récup)",
     "mode": "debutant", "play": "debutant", "flavor": "nager", "vol": "deb"},
    {"sheet": "02 Nager crawl", "title": "Nager · intermédiaire · crawl",
     "mode": "int_crawl", "play": "nager", "flavor": "nager", "vol": "int"},
    {"sheet": "03 Nager 4 nages", "title": "Nager · 4 nages (intermédiaire Oui + Avancé)",
     "mode": "four_n", "play": "four_n", "flavor": "nager", "vol": "mix"},
    {"sheet": "04 XS-Sprint deb crawl", "title": "XS / Sprint · débutant · crawl (+ 25 m dos/brasse récup)",
     "mode": "debutant", "play": "debutant", "flavor": "tri_short", "vol": "deb"},
    {"sheet": "05 XS-Sprint crawl", "title": "XS / Sprint · intermédiaire · crawl",
     "mode": "int_crawl", "play": "short", "flavor": "tri_short", "vol": "int"},
    {"sheet": "06 XS-Sprint 4 nages", "title": "XS / Sprint · 4 nages (intermédiaire Oui + Avancé)",
     "mode": "four_n", "play": "short", "flavor": "tri_short", "vol": "mix"},
    {"sheet": "07 Oly-Half-Full crawl", "title": "Olympique / Half / Full · intermédiaire · crawl",
     "mode": "int_crawl", "play": "long", "flavor": "tri_long", "vol": "int_long"},
    {"sheet": "08 Oly-Half-Full 4 nages", "title": "Olympique / Half / Full · 4 nages (intermédiaire Oui + Avancé)",
     "mode": "four_n", "play": "long", "flavor": "tri_long", "vol": "mix_long"},
    {"sheet": "09 OW courte deb crawl", "title": "Eau libre courte · débutant · crawl (+ 25 m dos/brasse récup)",
     "mode": "debutant", "play": "debutant", "flavor": "ow_short", "vol": "deb"},
    {"sheet": "10 OW courte crawl", "title": "Eau libre courte · intermédiaire · crawl",
     "mode": "int_crawl", "play": "ow_short", "flavor": "ow_short", "vol": "int"},
    {"sheet": "11 OW courte 4 nages", "title": "Eau libre courte · 4 nages (intermédiaire Oui + Avancé)",
     "mode": "four_n", "play": "ow_short", "flavor": "ow_short", "vol": "mix"},
    {"sheet": "12 OW moy-long crawl", "title": "Eau libre moyenne / longue · intermédiaire · crawl",
     "mode": "int_crawl", "play": "ow_long", "flavor": "ow_long", "vol": "int_long"},
    {"sheet": "13 OW moy-long 4 nages", "title": "Eau libre moyenne / longue · 4 nages (intermédiaire Oui + Avancé)",
     "mode": "four_n", "play": "ow_long", "flavor": "ow_long", "vol": "mix_long"},
]

METHODS = ["Arthur", "USA", "Canada", "Hongrie", "Italie"]


def method_of(i: int) -> str:
    return METHODS[i % 5]


def bande_of(fam: dict, i: int) -> str:
    if fam["mode"] == "debutant":
        return "débutant"
    if fam["mode"] == "int_crawl":
        return "intermédiaire"
    return "intermédiaire" if i % 2 == 0 else "avancé"


def vite(flavor: str) -> str:
    if flavor == "tri_short":
        return "vite, allure course"
    if flavor == "tri_long":
        return "vite, économie"
    if flavor.startswith("ow"):
        return "vite + sighting"
    return "vite"


def endu(flavor: str) -> str:
    if flavor == "tri_short":
        return "allure triathlon"
    if flavor == "tri_long":
        return "économie de nage"
    if flavor == "ow_short":
        return "crawl + sighting"
    if flavor == "ow_long":
        return "endurance orientation"
    return "allure régulière"


def rel(i: int) -> tuple[str, str]:
    return ("dos", "brasse") if i % 2 == 0 else ("brasse", "dos")


def rest_of(i: int) -> int:
    return (15, 20, 25, 30, 40)[i % 5]


def unique_marker(i: int, flavor: str, four: bool = False) -> str:
    """Série unique pour i dans 0..99 : n×50 × repos × texte."""
    n50 = 2 + (i % 10)
    rest = (15, 20, 25, 30, 40)[(i // 10) % 5]
    texts = [f"crawl {endu(flavor)}", "4 nages souple" if four else "crawl souple"]
    return nx(n50, 50, texts[i // 50], rest)


def inject_marker(bloc: str, i: int, flavor: str, four: bool, need: int) -> str:
    marker = unique_marker(i, flavor, four)
    lines = [ln for ln in bloc.split("\n") if ln]
    while lines and meters_of(join([marker] + lines)) > need:
        lines.pop()
    return join(pad([marker] + lines, need, fill_of(i, flavor, four), rest_of(i)))


def fill_of(i: int, flavor: str, four: bool = False) -> str:
    e, a = endu(flavor), rel(i)[0]
    if four:
        opts = ["nage libre souple", "crawl souple", "4 nages souple", f"crawl {e}", "dos souple"]
    else:
        opts = [
            "nage libre souple",
            "crawl souple",
            f"crawl {e}",
            f"en alternant (75 m crawl souple et 25 m {a})",
            "crawl respiration 3 temps",
        ]
    return opts[i % len(opts)]


def total_for(fam: dict, role: str, i: int, bande: str, phase: str | None = None) -> int:
    long = fam["vol"] in ("int_long", "mix_long") or "long" in fam["vol"]
    if phase == "test":
        if bande == "débutant":
            return 1400 if role == "test_50_100" else 1500
        if bande == "intermédiaire":
            return (2000 if role == "test_50_100" else 2100) if long else (1800 if role == "test_50_100" else 1900)
        return (2800 if role == "test_50_100" else 3000) if long else (2400 if role == "test_50_100" else 2600)
    if phase == "deload":
        idx = int(role[-1]) - 1
        if bande == "débutant":
            return [1000, 1100, 1200, 1300][idx]
        if bande == "intermédiaire":
            return [1400, 1500, 1600, 1700][idx] if long else [1200, 1300, 1400, 1500][idx]
        return [2000, 2100, 2200, 2400][idx] if long else [1800, 1900, 2000, 2100][idx]
    if bande == "débutant":
        t = [1400, 1500, 1600, 1500, 1700, 1500, 1400, 1600, 1800, 1500][i % 10]
        return min(t, 1500) if role == "séance récupération" else t
    if bande == "intermédiaire":
        cyc = [1800, 2000, 2200, 2000, 2300, 2100, 1900, 2400, 2000, 2200]
        if long:
            cyc = [2000, 2200, 2400, 2200, 2600, 2400, 2000, 2500, 2300, 2400]
        t = cyc[i % 10]
        if role == "séance récupération":
            t = min(t, 1800)
        if role == "séance vitesse":
            t = min(t, 2200)
        return t
    cyc = [2800, 3000, 3200, 3600, 3000, 4000, 3400, 4300, 3800, 3000]
    if long:
        cyc = [3000, 3400, 3800, 4200, 3600, 4600, 4000, 4300, 3800, 3400]
    t = cyc[i % 10]
    if role == "séance récupération":
        t = min(max(t, 2400), 2800)
    if role == "séance vitesse":
        t = min(t, 3400)
    return t


def budgets_special(total: int, bande: str, phase: str) -> tuple[int, int, int]:
    if bande == "débutant":
        ech, rac = 300, 100
    elif phase == "test":
        if bande == "intermédiaire":
            ech, rac = 400, 100
        else:
            ech, rac = 600, 200
    else:
        if bande == "intermédiaire":
            ech, rac = 200, 200
        else:
            ech, rac = 400, 200
    bloc = total - ech - rac
    if bloc < 400:
        ech, rac = 200, 100
        bloc = total - ech - rac
    return ech, bloc, rac


# ── Débutant (Gold Arthur s1/s2/s3) ─────────────────────────────────────────

def deb_ech(i: int) -> str:
    a, _ = rel(i)
    return join([
        "100 m nage libre souple",
        "50 m crawl",
        f"50 m {a}",
        "100 m crawl (25 m {éducatif} + 25 m crawl)",
    ])  # 300


def deb_test_deload_bloc(role: str, need: int, i: int, flavor: str) -> str | None:
    a, b = rel(i)
    course = vite(flavor)
    if role == "test_50_100":
        lines = [
            nx(8, 50, f"crawl {course}", 40),
            "100 m nage libre souple",
            nx(4, 100, f"crawl {course}", 60),
            f"200 m en alternant (75 m crawl et 25 m {a})",
            "100 m crawl {matériel}",
        ]
        return join(pad(lines, need, fill_of(i, flavor), rest_of(i)))
    if role == "test_400":
        lines = [
            f"400 m crawl {course}",
            "100 m nage libre souple",
            nx(2, 100, "crawl {matériel}", 20),
            f"200 m en alternant (75 m crawl et 25 m {a})",
            f"200 m en alternant (75 m crawl et 25 m {b})",
        ]
        return join(pad(lines, need, fill_of(i, flavor), rest_of(i)))
    if role == "deload_1":
        lines = [
            nx(8, 25, "crawl très court", 20),
            f"200 m en alternant (75 m crawl souple et 25 m {a})",
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_2":
        lines = [
            nx(4, 25, "crawl très court", 15),
            "200 m crawl souple",
            f"200 m en alternant (75 m crawl souple et 25 m {b})",
            "100 m crawl {matériel}",
        ]
        return join(pad(lines, need, "crawl souple", 20))
    if role == "deload_3":
        lines = [
            nx(8, 25, "crawl très court", 20),
            f"200 m en alternant (50 m {a} et 50 m {b})",
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_4":
        lines = [
            nx(4, 25, "crawl très court", 20),
            nx(4, 100, "crawl souple", 20),
            "100 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    return None


def deb_bloc(role: str, need: int, i: int, flavor: str, method: str) -> str:
    special = deb_test_deload_bloc(role, need, i, flavor)
    if special is not None:
        return special
    a, b = rel(i)
    v, e = vite(flavor), endu(flavor)
    r = rest_of(i)
    if role == "séance récupération":
        lines = [
            f"200 m en alternant (75 m crawl souple et 25 m {a})",
            nx(4, 100, "crawl souple", r),
            "100 m nage libre souple",
            f"200 m en alternant (75 m crawl souple et 25 m {b})",
        ]
        return join(pad(lines, need, fill_of(i, flavor), r))
    if role == "séance jambes":
        lines = [
            f"200 m en alternant (75 m crawl et 25 m {a})",
            nx(4, 50, "jambes crawl avec planche", r),
            "100 m nage libre souple",
            f"200 m crawl {e}",
            nx(2, 100, "crawl {matériel}", r),
        ]
        return join(pad(lines, need, f"crawl {e}", r))
    if role == "séance vitesse":
        lines = [
            nx(3, 100, f"crawl (50 m {v} + 50 m lent)", 20),
            nx(2, 100, "crawl {matériel}", 20),
            "200 m nage libre souple",
            nx(2, 100, f"crawl (75 m lent + 25 m {v})", 20),
            nx(2, 100, "crawl {matériel}", 20),
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "séance technique" or method == "Italie":
        lines = [
            f"200 m en alternant (75 m crawl et 25 m {a})",
            nx(3, 100, "crawl (75 m respiration 3 temps + 25 m respiration 5 temps)", r),
            f"100 m en alternant (75 m crawl et 25 m {b})",
            "100 m nage libre souple",
            "100 m crawl {matériel}",
            nx(3, 100, "crawl (75 m respiration 3 temps + 25 m respiration 5 temps)", r),
        ]
        return join(pad(lines, need, "crawl souple", r))
    if method == "USA":
        lines = [
            f"200 m en alternant (75 m crawl {e} et 25 m {a})",
            nx(4, 50, f"crawl {v}", 25),
            "100 m nage libre souple",
            "200 m crawl {matériel}",
            f"200 m en alternant (75 m crawl et 25 m {b})",
        ]
        return join(pad(lines, need, "nage libre souple", r))
    if method == "Canada":
        lines = [
            f"200 m en alternant (75 m crawl et 25 m {a})",
            nx(6, 100, f"crawl {e}", r),
            "100 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", r))
    if method == "Hongrie":
        lines = [
            f"200 m en alternant (50 m {a} et 50 m {b})",
            f"200 m en alternant (75 m crawl {v} et 25 m lent)",
            "100 m nage libre souple",
            nx(4, 50, "jambes crawl avec planche", r),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, "nage libre souple", r))
    # Arthur s2
    lines = [
        f"200 m en alternant (75 m crawl et 25 m {a})",
        f"200 m en alternant (50 m {a} et 50 m {b})",
        f"200 m en alternant (75 m {v} et 25 m lent)",
        "100 m nage libre souple",
        "200 m crawl {matériel}",
        f"200 m en alternant (75 m crawl et 25 m {b})",
    ]
    return join(pad(lines, need, fill_of(i, flavor), r))


def session_deb(fam, i, role, bande, phase=None):
    total = total_for(fam, role, i, bande, phase)
    method = method_of(i)
    ech = deb_ech(i)
    rac = "100 m nage libre souple"
    bloc = deb_bloc(role, total - 300 - 100, i, fam["flavor"], method)
    if not str(role).startswith(("test", "deload")):
        bloc = inject_marker(bloc, i, fam["flavor"], False, total - 400)
    return row(i, phase, bande, total, ech, bloc, rac)


# ── Intermédiaire crawl ──────────────────────────────────────────────────────

def int_c_ech(ech_m: int, method: str, i: int) -> str:
    r = 15
    if ech_m == 200:
        return "200 m crawl souple"
    if ech_m == 400:
        return nx(4, 100, "crawl (75 m souple + 25 m progressif)", r)
    if ech_m == 500:
        return join([
            "200 m crawl souple",
            nx(6, 50, "crawl (25 m souple + 25 m un cran plus vite)", r),
        ])
    if method == "USA" and ech_m % 100 == 0:
        return nx(ech_m // 100, 100, "crawl souple", r)
    return join(["200 m crawl souple", nx((ech_m - 200) // 50, 50, "crawl progressif", r)])


def int_c_bloc(role: str, need: int, i: int, flavor: str, method: str) -> str:
    v, e = vite(flavor), endu(flavor)
    r = rest_of(i)
    if role == "test_50_100":
        lines = [
            nx(8, 50, f"crawl {v}", 40),
            "100 m nage libre souple",
            nx(6, 100, f"crawl {v}", 60),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, fill_of(i, flavor), r))
    if role == "test_400":
        lines = [
            f"400 m crawl {v}",
            "200 m nage libre souple",
            nx(4, 100, f"crawl {e}", 30),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, fill_of(i, flavor), r))
    if role == "deload_1":
        lines = [
            nx(8, 25, "crawl très court", 20),
            nx(6, 100, "crawl souple", 20),
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_2":
        lines = [
            nx(4, 25, "crawl très court", 15),
            "200 m crawl souple",
            nx(4, 100, f"crawl {e}", 20),
            "100 m crawl {matériel}",
        ]
        return join(pad(lines, need, "crawl souple", 20))
    if role == "deload_3":
        lines = [
            nx(8, 25, "crawl très court", 20),
            nx(4, 50, "jambes crawl avec planche", 15),
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_4":
        lines = [
            nx(6, 25, "crawl très court", 20),
            nx(4, 100, "crawl souple", 20),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "séance récupération":
        return join(pad([nx(6, 100, "crawl souple", 20), "200 m crawl souple"], need, "nage libre souple", 20))
    if role == "séance jambes":
        lines = [
            nx(8, 50, "jambes crawl avec planche", 20),
            nx(4, 100, f"crawl {e}", r),
            "100 m crawl souple",
            nx(4, 50, "crawl {matériel}", 20),
        ]
        return join(pad(lines, need, fill_of(i, flavor), 25))
    if role == "séance vitesse":
        lines = [
            nx(4, 100, f"crawl (25 m {v} + 75 m souple)", 60),
            nx(3, 200, f"crawl {e}", 30),
            nx(4, 100, "crawl (25 m sprint + 75 m souple)", 60),
        ]
        return join(pad(lines, need, "crawl souple", 20))
    if role == "séance technique" or method == "Italie":
        lines = [
            nx(4, 100, f"crawl {e}", 30),
            "100 m crawl souple",
            nx(6, 100, f"crawl (50 m moyen + 25 m {v} + 25 m souple)", 45),
        ]
        return join(pad(lines, need, "crawl souple", 20))
    if method in ("Canada", "Arthur"):
        n400 = 1 + (i % 3) if need >= 1400 else 1
        lines = [
            nx(4, 100, f"crawl {e}", 30),
            "100 m nage libre souple",
            nx(n400, 400, "crawl (75 m respiration 3 temps + 25 m respiration 5 temps)", 60),
        ]
        return join(pad(lines, need, fill_of(i, flavor), 25))
    lines = [nx(8, 100, f"crawl {e}", 20), "200 m crawl souple", nx(4, 50, f"crawl {v}", 30)]
    return join(pad(lines, need, fill_of(i, flavor), r))


def budgets_int(total: int, role: str, i: int) -> tuple[int, int, int]:
    """échauffement, bloc, rac."""
    if total >= 2200:
        ech, rac = 500, 200
    elif total >= 2000:
        ech, rac = 400, 300 if i % 3 == 1 else 200
    elif total >= 1800:
        ech, rac = 200, 200
    else:
        ech, rac = 200, 100
    if role == "séance récupération":
        ech, rac = min(ech, 400), 200
    if role == "séance vitesse":
        rac = 100
        ech = min(ech, 500)
    bloc = total - ech - rac
    if bloc < 800:
        ech, rac = 200, 100
        bloc = total - ech - rac
    return ech, bloc, rac


def rac_txt(meters: int, four: bool, style: str = "int") -> str:
    if four and meters == 200:
        return "100 m dos souple\n100 m crawl souple"
    if four and meters == 300:
        return "300 m crawl souple"
    if style == "av" and meters == 200:
        return "200 m nage au choix souple"
    if meters == 300:
        return "300 m nage libre souple"
    if meters == 100:
        return "100 m nage libre souple"
    return f"{meters} m nage libre souple"


def session_int_crawl(fam, i, role, bande, phase=None):
    total = total_for(fam, role, i, bande, phase)
    method = method_of(i)
    if phase in ("test", "deload"):
        ech_m, bloc_m, rac_m = budgets_special(total, bande, phase)
    else:
        ech_m, bloc_m, rac_m = budgets_int(total, role, i)
    ech = int_c_ech(ech_m, method, i)
    if "{éducatif}" not in ech:
        # éducatif dans le bloc d'échauffement : 100 m du budget ech
        if ech_m >= 300:
            core = ech_m - 100
            ech = join([int_c_ech(core, method, i) if core >= 200 else f"{core} m crawl souple",
                        "100 m crawl (25 m {éducatif} + 25 m crawl)"])
            if meters_of(ech) != ech_m:
                ech = join([f"{ech_m - 100} m crawl souple",
                            "100 m crawl (25 m {éducatif} + 25 m crawl)"])
        else:
            ech = join([f"{ech_m - 100} m crawl souple" if ech_m > 100 else "100 m crawl souple",
                        "100 m crawl (25 m {éducatif} + 25 m crawl)"]) if ech_m >= 200 else ech
    if meters_of(ech) != ech_m:
        ech = join([f"{ech_m - 100} m crawl souple",
                    "100 m crawl (25 m {éducatif} + 25 m crawl)"])
    bloc = int_c_bloc(role, bloc_m, i, fam["flavor"], method)
    if not str(role).startswith(("test", "deload")):
        bloc = inject_marker(bloc, i, fam["flavor"], False, bloc_m)
    return row(i, phase, bande, total, ech, bloc, rac_txt(rac_m, False))


# ── Intermédiaire 4 nages ────────────────────────────────────────────────────

def int_4_ech(ech_m: int, i: int) -> str:
    r = 15
    g = i % 3
    if g == 0 and ech_m >= 500:
        extra = ech_m - 500
        lines = [
            "200 m nage libre crawl ou 4 nages",
            nx(6, 50, "(25 m dos + 25 m crawl)", r),
        ]
        if extra:
            lines.append(f"{extra} m crawl souple")
        return join(lines)
    if g == 1:
        return nx(ech_m // 100, 100, "(75 m crawl + 25 m dos ou 4 nages)", r) if ech_m % 100 == 0 else f"{ech_m} m 4 nages souple"
    # Gold s3 pyramid 200+150+100+50 = 500
    extra = ech_m - 500
    lines = [
        "200 m crawl lent",
        "150 m dos lent",
        "100 m brasse lent",
        "50 m papillon lent",
    ]
    if extra > 0:
        lines.append(f"{extra} m crawl souple")
    if ech_m < 500:
        return join(["200 m crawl ou 4 nages souple",
                     nx((ech_m - 200) // 50, 50, "(25 m dos + 25 m crawl)", r)])
    return join(lines)


def int_4_bloc(role: str, need: int, i: int, flavor: str, method: str) -> str:
    v, e = vite(flavor), endu(flavor)
    r = rest_of(i)
    if role == "test_50_100":
        lines = [
            nx(8, 50, f"crawl {v}", 40),
            "100 m nage libre souple",
            nx(6, 100, f"crawl {v}", 60),
            nx(4, 50, "4 nages souple", 20),
        ]
        return join(pad(lines, need, fill_of(i, flavor, True), r))
    if role == "test_400":
        lines = [
            f"400 m crawl {v}",
            "100 m dos souple",
            nx(4, 100, "4 nages", 30),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, fill_of(i, flavor, True), r))
    if role == "deload_1":
        lines = [
            nx(8, 25, "crawl très court", 20),
            nx(4, 100, "4 nages souple", 20),
            "200 m dos souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_2":
        lines = [
            nx(4, 25, "4 nages très court", 15),
            nx(4, 100, "crawl souple", 20),
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, "crawl souple", 20))
    if role == "deload_3":
        lines = [
            nx(8, 25, "crawl très court", 20),
            nx(4, 50, "jambes 4 nages avec planche", 15),
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "deload_4":
        lines = [
            nx(6, 25, "nage au choix très court", 20),
            nx(4, 100, "4 nages souple", 20),
            "100 m crawl {matériel}",
        ]
        return join(pad(lines, need, "nage libre souple", 20))
    if role == "séance récupération":
        return join(pad([nx(4, 100, "4 nages souple", 20), "200 m dos souple", nx(4, 100, "crawl souple", 20)], need, "nage libre souple"))
    if role == "séance jambes":
        lines = [
            nx(4, 100, "4 nages", 30),
            nx(8, 50, "jambes 4 nages avec planche", 20),
            "100 m nage libre souple",
            nx(4, 100, f"spécialité crawl {e}", 30),
        ]
        return join(pad(lines, need, "crawl souple"))
    if role == "séance vitesse":
        lines = [
            nx(4, 100, f"crawl (25 m {v} + 75 m souple)", 60),
            nx(3, 300, "(50 m crawl + 50 m dos + 50 m brasse), sinon 4 nages par 100 m", 40),
            nx(4, 100, "1 nage par 100 m (25 m sprint + 75 m souple)", 60),
        ]
        return join(pad(lines, need, "nage libre souple"))
    if method == "Italie" or role == "séance technique":
        lines = [
            nx(4, 100, "4 nages", 30),
            "100 m dos souple",
            nx(6, 100, f"spécialité (50 m moyen + 25 m {v} + 25 m souple)", 45),
        ]
        return join(pad(lines, need, fill_of(i, flavor, True)))
    n400 = 1 + (i % 3) if need >= 1400 else 1
    lines = [
        nx(4, 100, "4 nages", 30),
        "100 m nage libre souple",
        nx(n400, 400, "crawl (75 m respiration 3 temps + 25 m respiration 5 temps)", 60),
    ]
    return join(pad(lines, need, fill_of(i, flavor, True)))


def session_int_4n(fam, i, role, bande, phase=None):
    total = total_for(fam, role, i, bande, phase)
    method = method_of(i)
    if phase in ("test", "deload"):
        ech_m, bloc_m, rac_m = budgets_special(total, bande, phase)
    else:
        ech_m, bloc_m, rac_m = budgets_int(total, role, i)
    if phase not in ("test", "deload") and i % 3 == 0 and total >= 1800 and ech_m != 500:
        d = 500 - ech_m
        if bloc_m - d >= 800:
            ech_m, bloc_m = 500, bloc_m - d
    ech = int_4_ech(ech_m, i)
    if "{éducatif}" not in ech:
        if ech_m >= 400:
            keep = ech_m - 200
            ech = join([int_4_ech(keep, i) if meters_of(int_4_ech(keep, i)) == keep else f"{keep} m 4 nages souple",
                        "4 × 50 m {éducatif} (4 nages), repos 15 s"])
            if meters_of(ech) != ech_m:
                ech = join([f"{ech_m - 200} m crawl ou 4 nages souple",
                            "4 × 50 m {éducatif} (4 nages), repos 15 s"])
        else:
            ech = join([f"{ech_m - 100} m crawl ou 4 nages souple",
                        "2 × 50 m {éducatif} (4 nages), repos 15 s"]) if ech_m >= 200 else ech
    if meters_of(ech) != ech_m:
        ech = join([f"{ech_m - 200} m crawl ou 4 nages souple",
                    "4 × 50 m {éducatif} (4 nages), repos 15 s"])
        if meters_of(ech) != ech_m:
            ech = join([f"{ech_m - 100} m 4 nages souple",
                        "2 × 50 m {éducatif} (4 nages), repos 15 s"])
    bloc = int_4_bloc(role, bloc_m, i, fam["flavor"], method)
    if not str(role).startswith(("test", "deload")):
        bloc = inject_marker(bloc, i, fam["flavor"], True, bloc_m)
    rac = rac_txt(rac_m, True)
    if meters_of(rac) != rac_m:
        rac = f"{rac_m} m nage libre souple"
    return row(i, phase, bande, total, ech, bloc, rac)


# ── Avancé 4 nages ───────────────────────────────────────────────────────────

def budgets_av(total: int, role: str) -> tuple[int, int, int]:
    if total >= 4000:
        ech, rac = 1200, 200
    elif total >= 3600:
        ech, rac = 1200, 300
    elif total >= 3000:
        ech, rac = 1100, 300
    else:
        ech, rac = 600, 200
    if role == "séance récupération":
        ech, rac = 600, 200
    if role == "séance vitesse":
        rac = 200
        ech = min(ech, 800)
    bloc = total - ech - rac
    if bloc < 1400:
        ech, rac = 600, 200
        bloc = total - ech - rac
    return ech, bloc, rac


def av_ech(ech_m: int, i: int) -> str:
    g = i % 3
    if g == 0:
        # 2x100 4n inv + 8x50 prog + 5x100 éducatif = 200+400+500 = 1100
        core = join([
            nx(2, 100, "4 nages inversés, lents"),
            nx(8, 50, "crawl progressif de 1 à 4", 20),
            nx(5, 100, "4 nages {éducatif}", 30),
        ])
        have = 1100
        if ech_m == have:
            return core
        if ech_m > have:
            return join([core, f"{ech_m - have} m crawl souple"])
        # smaller: drop éducatif size
        edu = ech_m - 600
        if edu >= 200 and edu % 100 == 0:
            return join([
                nx(2, 100, "4 nages inversés, lents"),
                nx(8, 50, "crawl progressif de 1 à 4", 20),
                nx(edu // 100, 100, "4 nages {éducatif}", 30),
            ])
        return join([nx(2, 100, "4 nages inversés, lents"), f"{ech_m - 200} m crawl progressif"])
    if g == 1:
        if ech_m >= 1200:
            extra = ech_m - 1200
            lines = [
                "400 m nage libre, lent",
                "400 m 2 nages par 50 m, lent",
                "400 m 4 nages inversés",
            ]
            if extra:
                lines.append(f"{extra} m crawl souple")
            return join(lines)
        return join(["400 m nage libre, lent", f"{ech_m - 400} m 2 nages par 50 m, lent"])
    # g==2 : 6x100 4n + 6x100 crawl virage + 200 jambes = 1400
    if ech_m >= 1400:
        extra = ech_m - 1400
        lines = [
            nx(6, 100, "4 nages (1 normal, 1 inversé)", 20),
            nx(6, 100, "crawl lent, accélération à chaque virage", 30),
            "200 m jambes (25 m ondulation + 25 m battements)",
        ]
        if extra:
            lines.append(f"{extra} m crawl souple")
        return join(lines)
    if ech_m >= 800:
        return join([
            nx(6, 100, "4 nages (1 normal, 1 inversé)", 20),
            f"{ech_m - 600} m crawl lent, accélération à chaque virage",
        ])
    return nx(ech_m // 100, 100, "4 nages (1 normal, 1 inversé)", 20)


def av_bloc(role: str, need: int, i: int, flavor: str, method: str) -> str:
    v, e = vite(flavor), endu(flavor)
    g = i % 3
    if role == "test_50_100":
        lines = [
            nx(8, 50, f"crawl {v}", 40),
            "200 m nage libre souple",
            nx(8, 100, f"crawl {v}", 60),
            nx(4, 50, "nage au choix (25 m sprint + 25 m souple)", 30),
        ]
        return join(pad(lines, need, fill_of(i, flavor, True)))
    if role == "test_400":
        lines = [
            f"400 m crawl {v}",
            "200 m nage libre souple",
            nx(4, 100, "4 nages", 30),
            nx(4, 100, f"spécialité crawl {e}", 30),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, fill_of(i, flavor, True)))
    if role == "deload_1":
        lines = [
            nx(8, 25, "crawl très court", 20),
            nx(8, 100, "4 nages souple", 20),
        ]
        return join(pad(lines, need, "nage libre souple"))
    if role == "deload_2":
        lines = [
            nx(4, 25, "4 nages très court", 15),
            "400 m crawl souple",
            nx(4, 100, "dos souple", 20),
        ]
        return join(pad(lines, need, "nage libre souple"))
    if role == "deload_3":
        lines = [
            nx(8, 25, "nage au choix très court", 20),
            nx(4, 50, "jambes 4 nages avec planche", 15),
            "400 m crawl souple",
        ]
        return join(pad(lines, need, "nage libre souple"))
    if role == "deload_4":
        lines = [
            nx(6, 25, "crawl très court", 20),
            nx(6, 100, "4 nages souple", 20),
            "200 m crawl {matériel}",
        ]
        return join(pad(lines, need, "nage libre souple"))
    if role == "séance récupération":
        return join(pad([nx(8, 100, "4 nages souple", 20), "400 m crawl souple", nx(4, 100, "dos souple", 20)], need, "nage libre souple"))
    if role == "séance jambes":
        n200 = 8 if need >= 2200 else (4 if need >= 1600 else 2)
        lines = [
            "200 m jambes (25 m ondulation + 25 m battements)",
            nx(n200, 200, "en alternant 1 crawl et 1 4 nages jambes avec planche", 60),
            "200 m nage libre souple",
        ]
        return join(pad(lines, need, f"crawl {e}"))
    if role == "séance vitesse":
        lines = [
            f"400 m crawl {e}",
            "2 × 50 m crawl sprint, départ toutes les 2 min",
            "200 m nage libre souple",
            nx(3, 200, f"crawl {e}", 40),
            "200 m nage libre souple",
            nx(8, 50, "nage au choix (25 m sprint + 25 m souple)", 30),
        ]
        return join(pad(lines, need, "nage libre souple"))
    if g == 0 or method == "Italie":
        lines = [
            nx(6, 100, "crawl ou nage au choix (50 m moyen + 25 m rapide + 25 m souple)", 30),
            "800 m crawl avec pull-buoy + plaquettes" if need >= 1600 else "400 m crawl avec pull-buoy + plaquettes",
            nx(4, 50, "4 nages", 15),
        ]
        return join(pad(lines, need, f"crawl {e}"))
    if g == 1 or method == "Canada":
        n200 = (8 if need >= 2800 else (4 if need >= 2000 else 2)) + (i % 2)
        lines = [
            "4 × 50 m papillon (1 lent, 1 moyen, 1 rapide, 1 souple), départ 1 min",
            "4 × 50 m dos (1 lent, 1 moyen, 1 rapide, 1 souple), départ 1 min",
            "4 × 50 m brasse (1 lent, 1 moyen, 1 rapide, 1 souple), départ 1 min",
            "4 × 50 m crawl (1 lent, 1 moyen, 1 rapide, 1 souple), départ 1 min",
            nx(n200, 200, "en alternant 1 crawl et 1 4 nages jambes", 60),
            "200 m nage libre souple",
            "600 m crawl (50 m respiration 3 temps + 25 m respiration 5 temps + 25 m respiration 7 temps)" if need >= 2200 else "200 m crawl (respiration 3 / 5 temps)",
        ]
        return join(pad(lines, need, f"crawl {e}"))
    n500 = 3 if need >= 2700 else (2 if need >= 2200 else 1)
    lines = [
        f"400 m crawl {e}",
        "2 × 50 m crawl sprint, départ toutes les 2 min",
        "200 m nage libre souple",
        nx(n500, 500, "crawl avec pull-buoy", 60),
        "200 m nage libre souple",
        nx(6, 50, "nage au choix (25 m sprint + 25 m souple)", 30),
    ]
    return join(pad(lines, need, "nage libre souple"))


def av_ech_with_edu(ech_m: int, i: int) -> str:
    ech = av_ech(ech_m, i)
    if "{éducatif}" in ech:
        return ech
    if ech_m >= 500:
        keep = ech_m - 500
        base = av_ech(keep, i) if keep >= 400 else f"{keep} m 4 nages lent"
        if meters_of(base) != keep:
            base = f"{keep} m 4 nages lent"
        return join([base, nx(5, 100, "4 nages {éducatif}", 30)])
    return join([f"{ech_m - 200} m 4 nages lent", "4 × 50 m {éducatif} (4 nages), repos 20 s"])


def session_av(fam, i, role, bande, phase=None):
    total = total_for(fam, role, i, bande, phase)
    method = method_of(i)
    if phase in ("test", "deload"):
        ech_m, bloc_m, rac_m = budgets_special(total, bande, phase)
    else:
        ech_m, bloc_m, rac_m = budgets_av(total, role)
    ech = av_ech_with_edu(ech_m, i)
    if meters_of(ech) != ech_m:
        leftover = ech_m - meters_of(ech)
        if leftover > 0:
            ech = join([ech, f"{leftover} m crawl souple"])
        elif leftover < 0:
            ech = join([f"{ech_m - 500} m 4 nages lent", nx(5, 100, "4 nages {éducatif}", 30)])
        if meters_of(ech) != ech_m:
            raise ValueError(f"ech av {meters_of(ech)}≠{ech_m}\n{ech}")
    bloc = av_bloc(role, bloc_m, i, fam["flavor"], method)
    if not str(role).startswith(("test", "deload")):
        bloc = inject_marker(bloc, i, fam["flavor"], True, bloc_m)
    rac = rac_txt(rac_m, True, "av")
    if i % 3 == 2 and rac_m == 200:
        rac = "100 m crawl souple\n100 m dos souple"
    if meters_of(rac) != rac_m:
        rac = f"{rac_m} m nage libre souple"
    return row(i, phase, bande, total, ech, bloc, rac)


# ── row / excel ──────────────────────────────────────────────────────────────

def row(i, phase, bande, total, ech, bloc, rac) -> dict:
    blob = f"{ech}\n{bloc}\n{rac}"
    if "{éducatif}" not in blob:
        raise ValueError(f"#{i+1} pas de {{éducatif}}")
    s = meters_of(ech) + meters_of(bloc) + meters_of(rac)
    if s != total:
        raise ValueError(
            f"#{i+1} {phase} {bande} {s}≠{total} ({meters_of(ech)}+{meters_of(bloc)}+{meters_of(rac)})\n"
            f"ECH\n{ech}\nBLOC\n{bloc}\nRAC\n{rac}"
        )
    if total % 100:
        raise ValueError(total)
    if "—" in blob or "mise en route" in blob or "retour au calme" in blob:
        raise ValueError(f"libellé interdit dans l'exo #{i+1}: {blob[:200]}")
    return {
        "n": i + 1, "phase": phase, "bande": bande, "total": total,
        "ech": ech, "bloc": bloc, "rac": rac,
    }


def session_sig(rec: dict) -> tuple[str, str, str]:
    return (rec["ech"], rec["bloc"], rec["rac"])


def uniquify(rec: dict, seen: set, i: int) -> dict:
    if session_sig(rec) not in seen:
        return rec
    bloc = rec["bloc"]
    pairs = [
        ("8 × 50 m ", "4 × 100 m "),
        ("4 × 100 m ", "2 × 200 m "),
        ("6 × 50 m ", "3 × 100 m "),
        ("6 × 100 m ", "3 × 200 m "),
        ("4 × 50 m ", "2 × 100 m "),
        ("8 × 100 m ", "4 × 200 m "),
        ("8 × 25 m ", "4 × 50 m "),
        ("nage libre souple", "crawl souple"),
        ("crawl souple", "nage libre souple"),
        ("4 nages souple", "crawl souple"),
    ]
    candidates = []
    for new_rest in (15, 20, 25, 30, 40, 45, 60):
        candidates.append(re.sub(r"repos \d+ s", f"repos {new_rest} s", bloc, count=1))
    for a, b in pairs:
        if a in bloc:
            candidates.append(bloc.replace(a, b, 1))
    easy = [
        "100 m nage libre souple",
        "100 m crawl souple",
        "100 m crawl (respiration 3 temps)",
        "100 m crawl (respiration 5 temps)",
        "100 m en alternant (75 m crawl souple et 25 m dos)",
        "100 m en alternant (75 m crawl souple et 25 m brasse)",
    ]
    for src in easy:
        if src in bloc:
            for dst in easy:
                if dst != src:
                    candidates.append(bloc.replace(src, dst, 1))
    for cand in candidates:
        if meters_of(cand) != meters_of(bloc):
            continue
        trial = {**rec, "bloc": cand}
        if session_sig(trial) not in seen:
            return trial
    raise ValueError(f"doublon #{i + 1} {rec['phase']} {rec['bande']}")


def build(fam, i, role, phase=None):
    pad.variant = i
    bande = bande_of(fam, i)
    if fam["mode"] == "debutant":
        return session_deb(fam, i, role, bande, phase)
    if fam["mode"] == "int_crawl":
        return session_int_crawl(fam, i, role, bande, phase)
    if bande == "intermédiaire":
        return session_int_4n(fam, i, role, bande, phase)
    return session_av(fam, i, role, bande, phase)


def style_header(ws, row_i, n):
    for col in range(1, n + 1):
        cell = ws.cell(row_i, col)
        cell.fill = fill_blue
        cell.font = font_h
        cell.alignment = center
        cell.border = thin


def write_index(wb):
    ws = wb.active
    ws.title = "00 Lire d'abord"
    ws["A1"] = "Catalogue Gold Arthur — 13 familles × 100"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28
    notes = [
        "Structure : échauffement → bloc de séance → retour au calme.",
        "Arbre : Débutant = crawl + 25 m dos/brasse récup. Intermédiaire = crawl OU 4 nages. Avancé = 4 nages.",
        "Nager : pas de phase. Triathlon et eau libre : n°1-2 test (T50+T100, puis T400) · n°3-6 deload · n°7-100 construction.",
        "Test = 50 m, 100 m et 400 m allure course. Deload = charge basse, volume faible, intensité très courte. Construction = volume qui tient.",
        "Trou {éducatif} = catalogue éducatifs. Trou {matériel} = planche / palmes / pull / plaquettes. Jamais pull + palmes le même jour.",
        "Totaux ×100. contrôle_somme = échauffement + bloc + retour au calme.",
    ]
    for idx, line in enumerate(notes, start=3):
        ws.cell(idx, 1, line).font = font_small
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
    ws["A10"] = "Feuille"
    ws["B10"] = "Famille"
    ws["C10"] = "Qui"
    ws["D10"] = "Séances"
    style_header(ws, 10, 4)
    who = {"debutant": "Débutant", "int_crawl": "Intermédiaire crawl", "four_n": "Intermédiaire 4 nages + Avancé"}
    for i, fam in enumerate(FAMILIES):
        r = 11 + i
        ws.cell(r, 1, fam["sheet"]).font = font_body
        ws.cell(r, 2, fam["title"]).font = font_body
        ws.cell(r, 3, who[fam["mode"]]).font = font_body
        ws.cell(r, 4, 100).font = font_body
        for c in range(1, 5):
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill_soft if i % 2 == 0 else fill_alt
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A11"


def write_family(wb, fam):
    event = is_event_family(fam)
    headers = HEADERS_EPREUVE if event else HEADERS_NAGER
    last_col = chr(ord("A") + len(headers) - 1)
    ws = wb.create_sheet(fam["sheet"][:31])
    ws["A1"] = fam["title"]
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 26
    ws["A2"] = (
        "échauffement · bloc · retour au calme  ·  2 tests (50+100, 400) · 4 deload · 94 construction  ·  {éducatif}  ·  {matériel}"
        if event
        else "échauffement · bloc · retour au calme  ·  {éducatif}  ·  {matériel}  ·  totaux ×100"
    )
    ws["A2"].font = font_small
    ws.merge_cells(f"A2:{last_col}2")
    for col, h in enumerate(headers, start=1):
        ws.cell(3, col, h)
    style_header(ws, 3, len(headers))
    seen = set()
    for i in range(100):
        phase, role = spec_for(fam, i)
        rec = uniquify(build(fam, i, role, phase), seen, i)
        seen.add(session_sig(rec))
        r = 4 + i
        if event:
            vals = [rec["n"], rec["phase"], rec["bande"], rec["total"],
                    rec["ech"], rec["bloc"], rec["rac"],
                    meters_of(rec["ech"]) + meters_of(rec["bloc"]) + meters_of(rec["rac"])]
            wrap_from = 5
        else:
            vals = [rec["n"], rec["bande"], rec["total"],
                    rec["ech"], rec["bloc"], rec["rac"],
                    meters_of(rec["ech"]) + meters_of(rec["bloc"]) + meters_of(rec["rac"])]
            wrap_from = 4
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.font = font_body
            cell.border = thin
            cell.alignment = wrap if c >= wrap_from else center
        if event and rec["phase"]:
            ws.cell(r, 2).fill = PHASE_FILL[rec["phase"]]
        ws.row_dimensions[r].height = 96
    if event:
        widths = {"A": 6, "B": 14, "C": 16, "D": 11, "E": 48, "F": 62, "G": 36, "H": 16}
    else:
        widths = {"A": 6, "B": 16, "C": 11, "D": 48, "E": 62, "F": 36, "G": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    last = 103
    table = Table(
        displayName="T" + re.sub(r"[^A-Za-z0-9]", "", fam["sheet"])[:20],
        ref=f"A3:{last_col}{last}",
    )
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{last}"
    ws.sheet_properties.tabColor = "7E57C2" if fam["mode"] == "four_n" else BLUE


def main():
    wb = Workbook()
    write_index(wb)
    phases = {}
    for fam in FAMILIES:
        write_family(wb, fam)
        if is_event_family(fam):
            for i in range(100):
                ph, _ = spec_for(fam, i)
                phases[ph] = phases.get(ph, 0) + 1
    wb.save(OUT)
    print(f"OK {OUT}")
    print("séances", 13 * 100)
    print("phases épreuve", phases)


if __name__ == "__main__":
    main()
