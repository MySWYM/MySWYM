#!/usr/bin/env python3
"""Génère le lexique Sheet ↔ app MySWYM — UNE feuille (à glisser dans Google Sheet).

Usage : python3 docs/coach-ligne/build_lexique_sheet.py
Sortie : docs/coach-ligne/lexique-sheet-myswym.xlsx
Onglet : « Lexique MySWYM »
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("lexique-sheet-myswym.xlsx")
SHEET_NAME = "Lexique MySWYM"

NAVY, BLUE = "0A162C", "006BFD"
LINE = "D5DEE8"
thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_section = PatternFill("solid", fgColor="1B3A6B")
fill_alt = PatternFill("solid", fgColor="F7FAFC")
fill_warn = PatternFill("solid", fgColor="FFF3E0")
font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_h = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11, color=NAVY)
font_bold = Font(name="Calibri", size=11, bold=True, color=NAVY)
wrap = Alignment(wrap_text=True, vertical="top")

COLS = 5  # Section | Sujet | Zone | Sens / règle | Détail / exemple


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row):
    headers = ["Section", "Sujet", "Zone", "Sens / règle", "Détail / exemple"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.fill = fill_blue
        cell.font = font_h
        cell.border = thin
        cell.alignment = wrap


def write_section(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    cell = ws.cell(row, 1, title)
    cell.fill = fill_section
    cell.font = font_section
    cell.alignment = Alignment(vertical="center", horizontal="left")
    for c in range(1, COLS + 1):
        ws.cell(row, c).border = thin
        if c > 1:
            ws.cell(row, c).fill = fill_section
    ws.row_dimensions[row].height = 22
    return row + 1


def write_row(ws, row, section, sujet, zone, sens, detail, *, warn=False, bold_sujet=False):
    vals = (section, sujet, zone, sens, detail)
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.font = font_bold if (c == 2 and bold_sujet) else font_body
        cell.border = thin
        cell.alignment = wrap
        if warn:
            cell.fill = fill_warn
        elif row % 2 == 0:
            cell.fill = fill_alt
    if len(str(detail or "")) > 80 or len(str(sens or "")) > 60:
        ws.row_dimensions[row].height = 42
    return row + 1


def build_rows():
    """Yields ('section_title', None) or ('data', tuple of 5)."""
    # --- Tokens ---
    yield ("section", "1 · Tokens {D:} {@:} {éducatif} {matériel}")
    tokens = [
        ("{D:facile}", "Z1", "Départ perso rythme très aisé", "Int/avancé + Premium + T100. Débutant → repos 30 s", "8 × 100 m crawl, {D:facile}"),
        ("{D:endurance}", "Z2", "Départ perso rythme tenable longtemps", "Idem", "8 × 100 m crawl, {D:endurance}"),
        ("{D:seuil}", "Z3", "Départ perso allure soutenue / course", "Idem", "8 × 100 m crawl, {D:seuil}"),
        ("{D:VO2}", "Z4", "Départ perso séries dures (récup incomplète)", "Idem — nage vite, repos assez court", "8 × 50 m crawl, {D:VO2}"),
        ("{D:sprint}", "Max", "Départ perso explosif (récup quasi complète)", "Idem — nage max, repos long", "8 × 25 m crawl, {D:sprint}"),
        ("{@:facile}", "Z1", "Plage de temps @mm:ss–mm:ss", "Mêmes règles que D", "8 × 100 m crawl, {@:facile}"),
        ("{@:endurance}", "Z2", "Plage allure endurance", "Idem", "8 × 100 m, {D:endurance} {@:endurance}"),
        ("{@:seuil}", "Z3", "Plage allure seuil", "Idem", "6 × 100 m, {D:seuil} {@:seuil}"),
        ("{@:VO2}", "Z4", "Plage allure VO2", "Idem", "8 × 50 m, {D:VO2} {@:VO2}"),
        ("{@:sprint}", "Max", "Plage allure sprint", "Idem", "8 × 25 m, {D:sprint} {@:sprint}"),
        ("{éducatif}", "—", "1 éducatif tiré (niveau + nage)", "Tous niveaux Sheet soft", "4 × 50 m {éducatif}, repos 15 s"),
        ("{éducatif_pap}", "—", "1 éducatif papillon", "Lignes 4 nages", "… {éducatif_pap} …"),
        ("{éducatif_dos}", "—", "1 éducatif dos", "Idem", "… {éducatif_dos} …"),
        ("{éducatif_brasse}", "—", "1 éducatif brasse", "Idem", "… {éducatif_brasse} …"),
        ("{éducatif_crawl}", "—", "1 éducatif crawl", "Idem", "… {éducatif_crawl} …"),
        ("{matériel}", "—", "Matos optionnel (fiche ∩ inventaire)", "Si rien → token retiré. Pas pull+palmes même ligne", "4 × 50 m {éducatif} {matériel}"),
    ]
    for t in tokens:
        # section, sujet, zone, sens+règle, exemple
        yield ("data", ("Tokens", t[0], t[1], f"{t[2]} — {t[3]}", t[4]))
    yield (
        "data_warn",
        (
            "Tokens",
            "ALIAS",
            "—",
            "souple/lent/z1→facile · moyen/z2→endurance · vite/rapide/course/z3→seuil · z4/vo2max→VO2 · max/abloc→sprint",
            "Sur onglets débutant (01, 04, 09) : ne pas mettre {D:} / {@:} (inutile).",
        ),
    )

    # --- Filières ---
    yield ("section", "2 · Filières (à quoi ça sert)")
    filieres = [
        ("facile", "Z1", "Tu peux parler, technique clean", "Base, récup active, volume sans casse", "Échauff / RAC / jours légers ; débutants surtout · récup large"),
        ("endurance", "Z2", "Rythme stable longtemps", "Moteur aérobie (tri XS/Sprint, fond)", "Inter+ ; cœur du volume · récup moyenne"),
        ("seuil", "Z3", "Dur mais tenable en série", "Allure course / rythme élevé", "Inter+ qui prépare une distance · récup serrée"),
        ("VO2", "Z4", "Dans le rouge, tu finis la rep", "Puissance aérobie max", "Confirmés ; 1 bloc qualité · récup incomplète"),
        ("sprint", "Max", "Court, explosif, max", "Vitesse, départ, coup de rein", "Doses courtes · récup quasi complète"),
    ]
    for f in filieres:
        yield ("data", ("Filières", f[0], f[1], f"{f[2]} — {f[3]}", f[4]))
    yield (
        "data_warn",
        (
            "Filières",
            "VO2 ≠ sprint",
            "—",
            "VO2 = douloureux et répété ; sprint = maximal et bien récupéré",
            "Marge D : VO2 +15 s · sprint +28 s (repos plus long → D souvent plus large).",
        ),
    )

    # --- Mots libres ---
    yield ("section", "3 · Mots libres (sans accolades)")
    mots = [
        ("souple", "UI", "Pastille Souple", "Récup — ≠ lent"),
        ("lent", "UI", "Pastille Lent", "Lent contrôlé (technique), pas récup"),
        ("moyen", "UI", "Pastille Moyen", "Rythme régulier (mot libre, pas token D)"),
        ("progressif", "UI", "Pastille Progressif", "Accélère sur la distance"),
        ("vite / rapide", "UI", "Pastille Vite", ""),
        ("à bloc / sprint", "UI", "Pastilles dédiées", "Mot libre ; pour D perso utilise {D:sprint}"),
        ("2+ allures", "UI", "Pastille Enchaînement", "Ex. lent progressif · souple moyen vite"),
        ("repos 30 s / R30\"", "UI", "Pastille R", "Pause fixe"),
        ("départ 1 min / D2'", "UI", "Pastille D fixe", "Ne pas mélanger avec {D:…} sur la même ligne"),
    ]
    for m in mots:
        yield ("data", ("Mots libres", m[0], m[1], m[2], m[3]))

    # --- Rédiger ---
    yield ("section", "4 · Rédiger une séance")
    rediger = [
        ("1 ligne = 1 exercice", "—", "Un retour à la ligne = une carte dans l’app", ""),
        ("Structure colonnes", "—", "échauffement · bloc · rac · total_m = somme", ""),
        ("Ordre d’une ligne", "—", "[volume] [nage] [allure(s)] [, récup ou {D:}] [, {@:}] [, éducatif/matos]", ""),
        ("Exemple simple", "—", "8 × 100 m crawl moyen, repos 30 s", ""),
        ("Exemple avec pace", "—", "8 × 100 m crawl, {D:endurance} {@:endurance}", ""),
        ("Exemple VO2", "—", "8 × 50 m crawl, {D:VO2} {@:VO2}", ""),
        ("Exemple sprint", "—", "8 × 25 m crawl sprint, {D:sprint}", ""),
        ("Éducatif", "—", "4 × 50 m {éducatif} {matériel}, repos 15 s", ""),
        ("4 nages + éducatifs", "—", "Écrire « 4 nages » + « éducatif(s) »", "→ 1 éducatif / nage (pap→dos→brasse→crawl)"),
        ("Débutant", "—", "Pas de {D:} / {@:}", "mots souple/moyen + repos … s"),
        ("Orthographe tokens", "—", "Exactement {D:endurance}", "pas d’espace bizarre ; VO2 OK en majuscules"),
        ("À éviter", "—", "Deux D sur la même ligne", "D2' fixe + {D:seuil} ; allure course pour débutants"),
    ]
    for r in rediger:
        yield ("data", ("Rédiger", r[0], r[1], r[2], r[3]))

    # --- Familles ---
    yield ("section", "5 · Familles onglets (soft-branch 01–13)")
    familles = [
        ("01 Nager deb crawl", "Débutant · Nager", "Non (débutant)"),
        ("02 Nager crawl", "Inter crawl · Nager", "Oui si Premium + T100"),
        ("03 Nager 4 nages", "Inter 4n / Avancé · Nager", "Oui si Premium + T100"),
        ("04 XS-Sprint deb crawl", "Débutant · XS/Sprint", "Non (débutant)"),
        ("05 XS-Sprint crawl", "Inter crawl · XS/Sprint", "Oui si Premium + T100"),
        ("06 XS-Sprint 4 nages", "Inter 4n / Avancé · XS/Sprint", "Oui si Premium + T100"),
        ("07 Oly-Half-Full crawl", "Inter crawl · Oly/Half/Full", "Oui si Premium + T100"),
        ("08 Oly-Half-Full 4 nages", "Inter 4n / Avancé · Oly/Half/Full", "Oui si Premium + T100"),
        ("09 OW courte deb crawl", "Débutant · OW courte", "Non (débutant)"),
        ("10 OW courte crawl", "Inter crawl · OW courte", "Oui si Premium + T100"),
        ("11 OW courte 4 nages", "Inter 4n / Avancé · OW courte", "Oui si Premium + T100"),
        ("12 OW moy-long crawl", "Inter crawl · OW moy/long", "Oui si Premium + T100"),
        ("13 OW moy-long 4 nages", "Inter 4n / Avancé · OW moy/long", "Oui si Premium + T100"),
    ]
    for f in familles:
        yield ("data", ("Familles", f[0], "—", f[1], f"Pace {{D:}}/{{@:}} : {f[2]}"))

    # --- Calcul D ---
    yield ("section", "6 · Calcul D (indicatif — code pace-placeholders.js)")
    calcul = [
        ("facile", "+25 s", "zone easy (Z1)", "Ex. T100=1:30 ×100 m ≈ D2'25\" – D2'30\""),
        ("endurance", "+15 s", "milieu easy / seuil", "≈ D2' – D2'10\""),
        ("seuil", "+12 s", "zone threshold", "≈ D1'50\" – D2'"),
        ("VO2", "+15 s (incomplète)", "proche / un cran sous sprint", "≈ D1'40\" – D1'50\""),
        ("sprint", "+28 s (complète)", "plus vite que VO2 (max court)", "≈ D1'50\" – D2'05\" (repos long)"),
    ]
    for c in calcul:
        yield ("data", ("Calcul D", c[0], c[1], c[2], c[3]))
    yield (
        "data_warn",
        (
            "Calcul D",
            "Formule",
            "—",
            "D = arrondi 5 s ( T100 × mult × (rep_m/100) + marge )",
            "Distance lue sur la ligne (8 × 50 m → 50). Débutant / sans T100 / non Premium → {D:} = « repos 30 s » ; {@:} retiré.",
        ),
    )

    # --- Notes ---
    yield ("section", "7 · Notes coach (pièges & colonnes Sheet)")
    notes = [
        (
            "Colonne Notes (Éducatifs)",
            "Libre — pas un filtre app",
            "L’app lit Nom, Nage, niveaux, utilité, comment, Matériel, Garder. Notes = ton mémo coach.",
        ),
        ("Garder = oui", "Seul filtre « actif »", "Garder ≠ oui → jamais tiré."),
        (
            "Matériel optionnel",
            "et/ou = alternatives",
            "Ne gate pas l’éducatif. {matériel} seulement si ∩ inventaire. Jamais pull + palmes même ligne.",
        ),
        ("Débutant", "Jamais {D:} / {@:}", "Onglets 01 / 04 / 09 : souple / moyen + repos … s."),
        ("Inter / Avancé + pace", "Premium + T100", "Sinon même fallback (repos / token retiré)."),
        ("D fixe vs {D:}", "Ne pas mélanger", "D2' (horloge) + {D:seuil} = conflit."),
        ("4 nages + éducatifs", "1 éducatif / nage", "« 4 nages » + « éducatif(s) » → pap→dos→brasse→crawl."),
        ("Anti-doublon éducatif", "Pas 2× d’affilée", "Dernier exclu ; variété ~5 derniers si pool OK."),
        (
            "Calendrier vers J",
            "S-6 allégée · S-7 test",
            "S0 course (max 2) · S-1 allégée · S-2→S-5 travail · couple dès S-6. Garde début plan : 2 sem. travail.",
        ),
        (
            "Planning app",
            "Pastille = progression",
            "« Cette semaine » suit les semaines validées (Semaine 2 → S-n−1), pas seulement le calendrier.",
        ),
        ("Soft-branch 01–13", "Pas de fallback composeur", "Diplômes = composeur (pas d’onglet Sheet)."),
        ("Réf. séance", "séance n° ≠ ligne Sheet", "Réf. 01-42 = onglet+ligne · séance n°6 = compteur nageur."),
        ("Regen Excel", "python3 docs/coach-ligne/build_lexique_sheet.py", "1 onglet « Lexique MySWYM » — à glisser / importer dans le Google Sheet."),
    ]
    for n in notes:
        yield ("data", ("Notes", n[0], "—", n[1], n[2]))


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    title = ws.cell(1, 1, "MySWYM — Lexique Google Sheet (1 feuille · à coller dans le classeur séances)")
    title.fill = fill_navy
    title.font = font_title
    title.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28

    write_header(ws, 2)
    row = 3
    for kind, payload in build_rows():
        if kind == "section":
            row = write_section(ws, row, payload)
        elif kind == "data_warn":
            row = write_row(ws, row, *payload, warn=True, bold_sujet=True)
        else:
            row = write_row(ws, row, *payload)

    autosize(ws, [14, 28, 12, 42, 58])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(COLS)}{row - 1}"

    wb.save(OUT)
    print(f"OK → {OUT} (onglet « {SHEET_NAME} », {row - 1} lignes)")


if __name__ == "__main__":
    main()
