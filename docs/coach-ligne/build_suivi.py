#!/usr/bin/env python3
"""Génère suivi-nageurs-coach.xlsx — suivi coach de ligne."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "/Users/arthurnoel/MySWYM/docs/coach-ligne/suivi-nageurs-coach.xlsx"

NAVY, BLUE = "0A162C", "006BFD"
SOFT, ALT, INPUT, FORMULA = "E8F1FF", "F7FAFC", "FFF8E1", "F4F7FB"
LINE, MUTED = "D5DEE8", "5B6B7C"
GREEN, RED, AMBER = "C8E6C9", "FFCDD2", "FFE0B2"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_soft = PatternFill("solid", fgColor=SOFT)
fill_input = PatternFill("solid", fgColor=INPUT)
fill_formula = PatternFill("solid", fgColor=FORMULA)
fill_alt = PatternFill("solid", fgColor=ALT)
fill_guide = PatternFill("solid", fgColor="F0F4F8")
fill_green = PatternFill("solid", fgColor=GREEN)
fill_red = PatternFill("solid", fgColor=RED)
fill_amber = PatternFill("solid", fgColor=AMBER)
fill_grey = PatternFill("solid", fgColor="ECEFF1")

font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_h = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_label = Font(name="Calibri", size=11, bold=True, color=NAVY)
font_body = Font(name="Calibri", size=11, color=NAVY)
font_small = Font(name="Calibri", size=10, color=NAVY)
font_muted = Font(name="Calibri", size=10, italic=True, color=MUTED)
font_h_on_blue = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)


def tip(text: str) -> Comment:
    c = Comment(text, "Coach")
    c.width = 280
    c.height = 88
    return c


def parse_temps(ref: str) -> str:
    return (
        f'IF({ref}="","",IF(ISNUMBER({ref}),IF({ref}<1,{ref}*86400,{ref}),'
        f'IF(ISNUMBER(FIND(":",{ref})),'
        f'VALUE(LEFT({ref},FIND(":",{ref})-1))*60+'
        f'VALUE(SUBSTITUTE(MID({ref},FIND(":",{ref})+1,20),",",".")),'
        f'VALUE(SUBSTITUTE({ref},",",".")))))'
    )


def fmt_temps(sec_ref: str) -> str:
    return f'IF({sec_ref}="","",INT({sec_ref}/60)&":"&TEXT(MOD({sec_ref},60),"00.00"))'


def set_widths(ws, mapping: dict) -> None:
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def header_row(ws, row: int, headers: list, fill=None) -> None:
    ws.row_dimensions[row].height = 30
    style = fill or fill_navy
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.font = font_h
        cell.fill = style
        cell.alignment = center
        cell.border = thin


def input_cell(ws, r, c, value=None):
    cell = ws.cell(r, c, value)
    cell.fill = fill_input
    cell.font = font_body
    cell.border = thin
    cell.alignment = center
    return cell


def formula_cell(ws, r, c, value, num_fmt=None, wrap=False):
    cell = ws.cell(r, c, value)
    cell.fill = fill_formula
    cell.font = font_small
    cell.border = thin
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def add_table(ws, name: str, ref: str) -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(tab)


def add_list_dv(ws, named: str, sqref: str, prompt: str) -> None:
    v = DataValidation(
        type="list",
        formula1=f"={named}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur hors liste",
        error="Choisis une valeur du menu.",
        promptTitle="Liste",
        prompt=prompt,
        showInputMessage=True,
    )
    v.add(sqref)
    ws.add_data_validation(v)


def last_sec(row: int, proto: str, nage: str) -> str:
    a = f"$A{row}"
    return (
        f'IF({a}="","",IFERROR(INDEX(FILTER(tblTests[temps_sec],'
        f'(tblTests[id_nageur]={a})*(tblTests[protocole]="{proto}")*'
        f'(tblTests[nage]={nage})*(tblTests[valide]="oui")*'
        f'(tblTests[date]=MAXIFS(tblTests[date],tblTests[id_nageur],{a},'
        f'tblTests[protocole],"{proto}",tblTests[nage],{nage},'
        f'tblTests[valide],"oui"))),1),""))'
    )


def prev_sec(row: int, proto: str, nage: str) -> str:
    a = f"$A{row}"
    return (
        f'IF({a}="","",IFERROR(INDEX(FILTER(tblTests[temps_sec],'
        f'(tblTests[id_nageur]={a})*(tblTests[protocole]="{proto}")*'
        f'(tblTests[nage]={nage})*(tblTests[valide]="oui")*'
        f'(tblTests[date]=IFERROR(LARGE(FILTER(tblTests[date],'
        f'(tblTests[id_nageur]={a})*(tblTests[protocole]="{proto}")*'
        f'(tblTests[nage]={nage})*(tblTests[valide]="oui")),2),0))),1),""))'
    )


def last_date(row: int, proto: str, nage: str) -> str:
    a = f"$A{row}"
    return (
        f'IF({a}="","",IFERROR(MAXIFS(tblTests[date],tblTests[id_nageur],{a},'
        f'tblTests[protocole],"{proto}",tblTests[nage],{nage},'
        f'tblTests[valide],"oui"),""))'
    )


def build_listes(wb: Workbook) -> dict:
    ws = wb.create_sheet("LISTES")
    ws.sheet_properties.tabColor = "9BB0C8"
    catalogs = {
        "A": ("niveau", ["decouverte", "regulier", "sportif", "performance"]),
        "B": ("statut_nageur", ["actif", "pause", "competition", "archive"]),
        "C": ("sexe", ["F", "H"]),
        "D": ("cadre", ["perso", "club"]),
        "E": ("bassin", ["25", "50"]),
        "F": ("oui_non", ["oui", "non"]),
        "G": ("discipline", [
            "course_piscine", "eau_libre", "triathlon", "diplome",
            "nager_progresser", "reprendre",
        ]),
        "H": ("nage", ["crawl", "dos", "brasse", "papillon", "4N"]),
        "I": ("priorite", ["A", "B"]),
        "J": ("statut_objectif", ["en_cours", "atteint", "rate", "reporte"]),
        "K": ("protocole", ["T50", "T100", "T200", "T400", "CSS", "course"]),
        "L": ("depart", ["plonge", "poussee", "eau"]),
        "M": ("contexte", ["test_dedie", "fin_seance", "competition"]),
        "N": ("forme", ["bonne", "moyenne", "fatigue", "malade"]),
        "O": ("bassin_lieu", ["25", "50", "OW"]),
        "P": ("theme_seance", [
            "technique", "endurance", "seuil", "vitesse",
            "eau_libre", "specifique", "recuperation", "test",
        ]),
        "Q": ("faite_seance", ["oui", "non", "annulee"]),
        "R": ("type_semaine", ["normale", "allegee", "test", "reference"]),
        "S": ("qualite", ["tenue", "limite", "casse"]),
    }
    ws.merge_cells("A1:S1")
    ws["A1"] = "Listes — tu peux ajouter des valeurs sous une colonne, puis élargir le nom défini."
    ws["A1"].font = font_muted
    keys = []
    for col, (name, values) in catalogs.items():
        ws[f"{col}2"] = name
        ws[f"{col}2"].font = font_h
        ws[f"{col}2"].fill = fill_navy
        ws[f"{col}2"].alignment = center
        for i, val in enumerate(values, start=3):
            ws[f"{col}{i}"] = val
        last = 2 + len(values)
        wb.defined_names.add(
            DefinedName(name=f"liste_{name}", attr_text=f"LISTES!${col}$3:${col}${last}")
        )
        keys.append(name)
    set_widths(ws, {c: 18 for c in "ABCDEFGHIJKLMNOPQRS"})
    ws.sheet_state = "hidden"
    return {k: f"liste_{k}" for k in keys}


def build_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("GUIDE", 0)
    ws.sheet_properties.tabColor = BLUE
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    set_widths(ws, {"A": 3, "B": 30, "C": 50, "D": 42, "E": 22})

    ws.merge_cells("B2:E2")
    ws["B2"] = "SUIVI NAGEURS — COACH DE LIGNE"
    ws["B2"].font = font_title
    ws["B2"].fill = fill_navy
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 34
    for col in "BCDE":
        ws[f"{col}2"].fill = fill_navy

    ws.merge_cells("B3:E3")
    ws["B3"] = (
        "Excel 365 (Mac ou Windows). Jaune = tu saisis. Gris = formule, ne pas écraser. "
        "Léa / Thomas / Inès sont des EXEMPLES — remplace-les par tes nageurs."
    )
    ws["B3"].font = font_muted
    ws["B3"].alignment = left
    ws.row_dimensions[3].height = 28

    blocks = [
        ("1. Logique",
         "1 ligne dans NAGEURS, 1 objectif A dans OBJECTIFS, 1 ligne par chrono dans TESTS "
         "(on n’écrase jamais un temps). Le DASHBOARD se calcule tout seul. "
         "EXPORT reprend les champs utiles pour MySWYM (T100, cible, date)."),
        ("2. Batterie A — vitesse (toutes les 4–5 semaines)",
         "Même échauffement à chaque fois (600–800 m). T50 à fond, récup 6–8 min, T100 à fond. "
         "Noter bassin (25/50), départ (plongé / poussée), nage. Pas de 400 le même jour."),
        ("3. Batterie B — aérobie (toutes les 6–8 semaines)",
         "T200 quasi max régulier, récup 10–15 min, T400. CSS = (T400 − T200) / 2 "
         "(allure seuil aux 100 m). Utile 400 / triathlon / eau libre. "
         "MySWYM n’utilise que le T100 ; le 400 est pour toi."),
        ("4. Saisie des temps",
         "1:12.50 ou 34.8 ou 0:34.80. Virgule française 1:12,50 acceptée. "
         "Splits 50 / 100 optionnels, précieux pour voir où ça casse."),
        ("5. Un seul objectif A",
         "Priorité A + statut en_cours = ce que le dashboard affiche. Un B possible. "
         "À 8–10 semaines : spécifique. À moins de 3 semaines : plus de test à fond."),
        ("6. Lire le diagnostic",
         "Drop 100 vs 2×50 > 10 % → il casse (lactique / pacing / technique). "
         "Drop 400 vs 4×100 > 12 % → manque d’aérobie. "
         "Homogène → allure spécifique vers la cible, pas un thème miracle."),
        ("7. Charge, deload, retest",
         "1 ligne par séance dans SEANCES (1 à 5 / semaine). CHARGE agrège par semaine. "
         "Deload : 3 semaines de charge puis 1 allégée (−30 % vol, RPE ≤ 6). "
         "Retest A (T50+T100) après 4–5 sem, Batterie B après 6–8 sem — jamais pendant un deload ni < 3 sem avant l’objectif."),
    ]
    r = 5
    for title, body in blocks:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws[f"B{r}"] = title
        ws[f"B{r}"].font = font_h_on_blue
        ws[f"B{r}"].fill = fill_blue
        ws[f"B{r}"].alignment = left
        for col in "BCD":
            ws[f"{col}{r}"].fill = fill_blue
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=4)
        ws[f"B{r}"] = body
        ws[f"B{r}"].font = font_body
        ws[f"B{r}"].alignment = left_top
        ws[f"B{r}"].fill = fill_guide
        ws.row_dimensions[r].height = 24
        ws.row_dimensions[r + 1].height = 24
        r += 3

    ws["B33"] = "Onglet"
    ws["C33"] = "Rôle"
    ws["B33"].font = font_h
    ws["C33"].font = font_h
    ws["B33"].fill = fill_navy
    ws["C33"].fill = fill_navy
    ws.merge_cells("C33:D33")
    tabs = [
        ("DASHBOARD", "Lecture dimanche : temps + deload + retest."),
        ("NAGEURS", "1 ligne = 1 personne. id stable (N001…)."),
        ("OBJECTIFS", "1 ligne = 1 course / diplôme."),
        ("TESTS", "1 ligne = 1 chrono. Cœur du suivi."),
        ("SEANCES", "1 ligne = 1 séance (1 à 5 / semaine)."),
        ("CHARGE", "1 ligne = 1 nageur × 1 semaine. Alertes auto."),
        ("EVOLUTION", "Courbe T100 d’exemple (Léa)."),
        ("EXPORT", "Colonnes app — ne pas saisir."),
    ]
    for i, (name, desc) in enumerate(tabs):
        row = 34 + i
        ws[f"B{row}"] = name
        ws[f"B{row}"].font = font_label
        ws[f"B{row}"].fill = fill_soft
        ws[f"C{row}"] = desc
        ws[f"C{row}"].font = font_body
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    ws["B43"] = "Jaune = saisie"
    ws["B43"].fill = fill_input
    ws["B43"].font = font_label
    ws["C43"] = "Gris = formule"
    ws["C43"].fill = fill_formula
    ws["C43"].font = font_label
    ws.freeze_panes = "B5"


def build_nageurs(wb: Workbook, lists: dict) -> None:
    ws = wb.create_sheet("NAGEURS")
    ws.sheet_properties.tabColor = BLUE
    headers = [
        "id", "prenom", "nom", "annee_nais", "sexe", "tel", "mail", "cadre",
        "bassin", "seances_sem", "niveau", "materiel", "blessure",
        "note_contrainte", "statut", "notes",
    ]
    header_row(ws, 1, headers)
    ws["A1"].comment = tip("N001, N002… Ne jamais changer un id déjà utilisé dans TESTS.")
    ws["K1"].comment = tip("decouverte / regulier / sportif / performance — mêmes libellés que MySWYM.")
    ws["J1"].comment = tip("Séances réellement tenables, pas le souhait.")

    rows = [
        ["N001", "Lea", "Martin", 1998, "F", "", "", "perso", 25, 3,
         "sportif", "pull, palmes", "non", "", "actif", "EXEMPLE — 100 NL club"],
        ["N002", "Thomas", "Bernard", 1992, "H", "", "", "perso", 50, 5,
         "performance", "plaquettes, tuba", "non", "", "actif", "EXEMPLE — 400 NL"],
        ["N003", "Ines", "Moreau", 2001, "F", "", "", "perso", 25, 3,
         "regulier", "pull", "non", "Genou sensible", "actif", "EXEMPLE — triathlon 1900"],
    ]
    last = 23
    for r, data in enumerate(rows, 2):
        for c, val in enumerate(data, 1):
            cell = input_cell(ws, r, c, val)
            if c in (12, 14, 16):
                cell.alignment = left
        ws.row_dimensions[r].height = 22
    for r in range(5, last + 1):
        for c in range(1, 17):
            input_cell(ws, r, c)
        ws.row_dimensions[r].height = 20

    add_table(ws, "tblNageurs", f"A1:P{last}")
    set_widths(ws, {
        "A": 10, "B": 14, "C": 14, "D": 12, "E": 8, "F": 14, "G": 24,
        "H": 10, "I": 10, "J": 12, "K": 14, "L": 22, "M": 10, "N": 22,
        "O": 12, "P": 36,
    })
    ws.freeze_panes = "D2"
    add_list_dv(ws, lists["sexe"], "E2:E200", "Sexe")
    add_list_dv(ws, lists["cadre"], "H2:H200", "Cadre")
    add_list_dv(ws, lists["bassin"], "I2:I200", "Bassin habituel")
    add_list_dv(ws, lists["niveau"], "K2:K200", "Niveau")
    add_list_dv(ws, lists["oui_non"], "M2:M200", "Blessure en cours ?")
    add_list_dv(ws, lists["statut_nageur"], "O2:O200", "Statut")


def build_objectifs(wb: Workbook, lists: dict) -> None:
    ws = wb.create_sheet("OBJECTIFS")
    ws.sheet_properties.tabColor = "3D8FFF"
    headers = [
        "id_objectif", "id_nageur", "discipline", "nage", "distance_m",
        "date_objectif", "lieu", "bassin_lieu", "temps_cible", "temps_cible_sec",
        "priorite", "statut", "notes", "cle_dashboard",
    ]
    header_row(ws, 1, headers)
    ws["K1"].comment = tip("Un seul A en_cours par nageur — c’est lui que le dashboard affiche.")
    ws["I1"].comment = tip("1:12.00 ou 5:05.00 ou 32:00")

    examples = [
        ["O001", "N001", "course_piscine", "crawl", 100, date(2026, 10, 18),
         "Meeting club — Lyon", "25", "1:12.00", "A", "en_cours", "Premier 100 officiel"],
        ["O002", "N002", "course_piscine", "crawl", 400, date(2026, 12, 6),
         "Championnats maitres — Paris", "50", "5:05.00", "A", "en_cours", ""],
        ["O003", "N003", "triathlon", "crawl", 1900, date(2026, 9, 13),
         "Triathlon M — Annecy", "OW", "32:00", "A", "en_cours", "Eau libre, pas de mur"],
    ]
    last = 23
    input_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13]
    for r, data in enumerate(examples, 2):
        input_cell(ws, r, 1, data[0])
        input_cell(ws, r, 2, data[1])
        input_cell(ws, r, 3, data[2])
        input_cell(ws, r, 4, data[3])
        input_cell(ws, r, 5, data[4])
        input_cell(ws, r, 6, data[5]).number_format = "YYYY-MM-DD"
        input_cell(ws, r, 7, data[6]).alignment = left
        input_cell(ws, r, 8, data[7])
        input_cell(ws, r, 9, data[8])
        input_cell(ws, r, 11, data[9])
        input_cell(ws, r, 12, data[10])
        input_cell(ws, r, 13, data[11]).alignment = left
        formula_cell(ws, r, 10, f"={parse_temps(f'I{r}')}", "0.00")
        formula_cell(ws, r, 14, f'=IF(AND(K{r}="A",L{r}="en_cours"),B{r},"")')
        ws.row_dimensions[r].height = 22
    for r in range(5, last + 1):
        for c in input_cols:
            input_cell(ws, r, c)
        ws.cell(r, 6).number_format = "YYYY-MM-DD"
        formula_cell(ws, r, 10, f"={parse_temps(f'I{r}')}", "0.00")
        formula_cell(ws, r, 14, f'=IF(AND(K{r}="A",L{r}="en_cours"),B{r},"")')

    add_table(ws, "tblObjectifs", f"A1:N{last}")
    set_widths(ws, {
        "A": 12, "B": 12, "C": 16, "D": 12, "E": 12, "F": 14, "G": 32,
        "H": 12, "I": 14, "J": 16, "K": 10, "L": 12, "M": 28, "N": 16,
    })
    ws.freeze_panes = "C2"
    add_list_dv(ws, lists["discipline"], "C2:C200", "Discipline")
    add_list_dv(ws, lists["nage"], "D2:D200", "Nage")
    add_list_dv(ws, lists["bassin_lieu"], "H2:H200", "Bassin du lieu")
    add_list_dv(ws, lists["priorite"], "K2:K200", "Priorité")
    add_list_dv(ws, lists["statut_objectif"], "L2:L200", "Statut")


def write_test_formulas(ws, r: int) -> None:
    formula_cell(ws, r, 11, f"={parse_temps(f'J{r}')}", "0.00")
    formula_cell(
        ws, r, 12,
        f'=IF(I{r}<>"",I{r},SWITCH(D{r},"T50",50,"T100",100,"T200",200,"T400",400,"CSS",400,""))',
    )
    formula_cell(ws, r, 13, f'=IF(OR(K{r}="",L{r}=""),"",K{r}*100/L{r})', "0.00")


def build_tests(wb: Workbook, lists: dict) -> None:
    ws = wb.create_sheet("TESTS")
    ws.sheet_properties.tabColor = BLUE
    headers = [
        "id_test", "id_nageur", "date", "protocole", "nage", "bassin", "depart",
        "contexte", "distance_saisie", "temps", "temps_sec", "distance_m",
        "allure_100_sec", "split_50", "split_100", "rpe", "forme", "valide", "notes",
    ]
    header_row(ws, 1, headers)
    ws["D1"].comment = tip("T50 T100 T200 T400 CSS course. Ne pas tout faire le même jour.")
    ws["I1"].comment = tip("Vide sauf protocole course (ex. 800) : la distance se déduit sinon.")
    ws["R1"].comment = tip("non = chrono douteux, gardé mais ignoré du dashboard.")
    ws["J1"].comment = tip("1:12.50 ou 34.8 — virgule OK.")

    examples = [
        ["T001", "N001", date(2026, 3, 15), "T50", "crawl", 25, "plonge", "test_dedie", None, "0:34.80", 8, "bonne", "oui", "Premier test saison"],
        ["T002", "N001", date(2026, 3, 15), "T100", "crawl", 25, "plonge", "test_dedie", None, "1:16.20", 9, "bonne", "oui", ""],
        ["T003", "N001", date(2026, 4, 12), "T400", "crawl", 25, "poussee", "test_dedie", None, "5:42.00", 8, "moyenne", "oui", "Batterie B"],
        ["T004", "N001", date(2026, 4, 12), "T200", "crawl", 25, "poussee", "test_dedie", None, "2:44.00", 8, "moyenne", "oui", ""],
        ["T005", "N001", date(2026, 5, 10), "T50", "crawl", 25, "plonge", "test_dedie", None, "0:34.10", 8, "bonne", "oui", ""],
        ["T006", "N001", date(2026, 5, 10), "T100", "crawl", 25, "plonge", "test_dedie", None, "1:14.80", 9, "bonne", "oui", ""],
        ["T007", "N001", date(2026, 7, 20), "T50", "crawl", 25, "plonge", "test_dedie", None, "0:33.60", 8, "bonne", "oui", ""],
        ["T008", "N001", date(2026, 7, 20), "T100", "crawl", 25, "plonge", "test_dedie", None, "1:13.40", 9, "bonne", "oui", "PB 100"],
        ["T009", "N001", date(2026, 7, 20), "T200", "crawl", 25, "poussee", "test_dedie", None, "2:42.00", 8, "bonne", "oui", "T200 après récup"],
        ["T010", "N002", date(2026, 2, 8), "T50", "crawl", 50, "plonge", "test_dedie", None, "0:28.80", 8, "bonne", "oui", ""],
        ["T011", "N002", date(2026, 2, 8), "T100", "crawl", 50, "plonge", "test_dedie", None, "1:04.20", 9, "bonne", "oui", ""],
        ["T012", "N002", date(2026, 3, 22), "T200", "crawl", 50, "poussee", "test_dedie", None, "2:20.00", 8, "bonne", "oui", ""],
        ["T013", "N002", date(2026, 3, 22), "T400", "crawl", 50, "poussee", "test_dedie", None, "4:58.00", 9, "bonne", "oui", ""],
        ["T014", "N002", date(2026, 6, 14), "T50", "crawl", 50, "plonge", "test_dedie", None, "0:28.10", 8, "bonne", "oui", ""],
        ["T015", "N002", date(2026, 6, 14), "T100", "crawl", 50, "plonge", "test_dedie", None, "1:02.40", 9, "bonne", "oui", ""],
        ["T016", "N002", date(2026, 6, 28), "T200", "crawl", 50, "poussee", "test_dedie", None, "2:16.80", 8, "bonne", "oui", ""],
        ["T017", "N002", date(2026, 6, 28), "T400", "crawl", 50, "poussee", "test_dedie", None, "4:52.00", 9, "bonne", "oui", "PB 400"],
        ["T018", "N003", date(2026, 4, 5), "T100", "crawl", 25, "poussee", "test_dedie", None, "1:32.00", 8, "moyenne", "oui", ""],
        ["T019", "N003", date(2026, 4, 5), "T400", "crawl", 25, "poussee", "test_dedie", None, "6:40.00", 8, "moyenne", "oui", ""],
        ["T020", "N003", date(2026, 6, 21), "T100", "crawl", 25, "poussee", "test_dedie", None, "1:28.00", 8, "bonne", "oui", ""],
        ["T021", "N003", date(2026, 6, 21), "T200", "crawl", 25, "poussee", "test_dedie", None, "3:05.00", 8, "bonne", "oui", ""],
        ["T022", "N003", date(2026, 6, 21), "T400", "crawl", 25, "poussee", "test_dedie", None, "6:20.00", 8, "bonne", "oui", ""],
        ["T023", "N003", date(2026, 8, 2), "T100", "crawl", 25, "poussee", "test_dedie", None, "1:26.50", 7, "bonne", "oui", "Avant spécifique"],
    ]
    last = 80
    input_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14, 15, 16, 17, 18, 19]
    for r, row in enumerate(examples, 2):
        input_cell(ws, r, 1, row[0])
        input_cell(ws, r, 2, row[1])
        input_cell(ws, r, 3, row[2]).number_format = "YYYY-MM-DD"
        input_cell(ws, r, 4, row[3])
        input_cell(ws, r, 5, row[4])
        input_cell(ws, r, 6, row[5])
        input_cell(ws, r, 7, row[6])
        input_cell(ws, r, 8, row[7])
        input_cell(ws, r, 9, row[8])
        input_cell(ws, r, 10, row[9])
        input_cell(ws, r, 14)
        input_cell(ws, r, 15)
        input_cell(ws, r, 16, row[10])
        input_cell(ws, r, 17, row[11])
        input_cell(ws, r, 18, row[12])
        input_cell(ws, r, 19, row[13]).alignment = left
        write_test_formulas(ws, r)
        ws.row_dimensions[r].height = 20
    for r in range(len(examples) + 2, last + 1):
        for c in input_cols:
            input_cell(ws, r, c)
        ws.cell(r, 3).number_format = "YYYY-MM-DD"
        write_test_formulas(ws, r)

    add_table(ws, "tblTests", f"A1:S{last}")
    set_widths(ws, {
        "A": 10, "B": 12, "C": 12, "D": 12, "E": 12, "F": 10, "G": 12,
        "H": 14, "I": 16, "J": 12, "K": 12, "L": 12, "M": 14, "N": 12,
        "O": 12, "P": 8, "Q": 12, "R": 10, "S": 36,
    })
    ws.freeze_panes = "D2"
    add_list_dv(ws, lists["protocole"], "D2:D200", "Protocole")
    add_list_dv(ws, lists["nage"], "E2:E200", "Nage")
    add_list_dv(ws, lists["bassin"], "F2:F200", "Bassin")
    add_list_dv(ws, lists["depart"], "G2:G200", "Départ")
    add_list_dv(ws, lists["contexte"], "H2:H200", "Contexte")
    add_list_dv(ws, lists["forme"], "Q2:Q200", "Forme")
    add_list_dv(ws, lists["oui_non"], "R2:R200", "Chrono exploitable ?")
    rpe = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    rpe.add("P2:P200")
    ws.add_data_validation(rpe)


def _week_sessions(nid, monday, week_type, items):
    """items: (day_offset, theme, vol, rpe, faite, qualite, lien_test, notes)"""
    rows = []
    for off, theme, vol, rpe, faite, qualite, lien, notes in items:
        rows.append([
            nid, monday + timedelta(days=off), theme, vol, rpe,
            faite, week_type, qualite, lien or "", notes,
        ])
    return rows


def build_seances(wb: Workbook, lists: dict) -> None:
    ws = wb.create_sheet("SEANCES")
    ws.sheet_properties.tabColor = "9BB0C8"
    headers = [
        "id_nageur", "date", "theme", "volume_m", "rpe", "faite",
        "type_semaine", "qualite", "lien_test", "notes",
        "lundi", "semaine_iso", "cle",
    ]
    header_row(ws, 1, headers)
    ws["G1"].comment = tip("À poser sur la 1re séance de la semaine. CHARGE reprend ce type.")
    ws["H1"].comment = tip("tenue = OK · limite = tout juste · casse = séance ratée / arrêt.")
    ws["K1"].comment = tip("Lundi ISO — formule, ne pas saisir.")

    examples = []
    # Léa N001 — 3×/sem, deload W32 (20 jul), 3 charges ensuite → deload dû W36
    examples += _week_sessions("N001", date(2026, 6, 29), "normale", [
        (1, "technique", 1800, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2200, 6, "oui", "tenue", "", ""),
        (5, "seuil", 2000, 7, "oui", "tenue", "", "10x100"),
    ])
    examples += _week_sessions("N001", date(2026, 7, 6), "normale", [
        (1, "technique", 2000, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2400, 6, "oui", "tenue", "", ""),
        (5, "vitesse", 2100, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N001", date(2026, 7, 13), "normale", [
        (1, "endurance", 2200, 6, "oui", "tenue", "", ""),
        (3, "seuil", 2500, 7, "oui", "limite", "", ""),
        (5, "specifique", 2300, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N001", date(2026, 7, 20), "allegee", [
        (0, "recuperation", 1400, 4, "oui", "tenue", "", "Deload"),
        (1, "test", 1600, 8, "oui", "tenue", "T100", "Batterie A"),
    ])
    examples += _week_sessions("N001", date(2026, 7, 27), "normale", [
        (1, "technique", 2000, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2300, 6, "oui", "tenue", "", ""),
        (5, "seuil", 2100, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N001", date(2026, 8, 3), "normale", [
        (1, "technique", 2200, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2500, 6, "oui", "tenue", "", ""),
        (5, "specifique", 2200, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N001", date(2026, 8, 10), "normale", [
        (1, "endurance", 2400, 6, "oui", "tenue", "", ""),
        (3, "seuil", 2600, 7, "oui", "limite", "", ""),
        (5, "vitesse", 2400, 8, "oui", "limite", "", "3e sem de charge"),
    ])
    examples += _week_sessions("N001", date(2026, 8, 17), "normale", [
        (1, "technique", 2300, 6, "oui", "tenue", "", ""),
        (3, "seuil", 2500, 7, "oui", "tenue", "", "10x100 tenus 1:25"),
        (5, "specifique", 2400, 7, "oui", "tenue", "", ""),
    ])
    # Thomas N002 — 5×/sem, deload 13 jul et 10 aout
    examples += _week_sessions("N002", date(2026, 6, 29), "normale", [
        (0, "endurance", 3200, 6, "oui", "tenue", "", ""),
        (1, "technique", 2800, 5, "oui", "tenue", "", ""),
        (2, "seuil", 3400, 7, "oui", "tenue", "", ""),
        (3, "vitesse", 3000, 8, "oui", "tenue", "", ""),
        (4, "specifique", 3600, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 7, 6), "normale", [
        (0, "endurance", 3400, 6, "oui", "tenue", "", ""),
        (1, "technique", 3000, 5, "oui", "tenue", "", ""),
        (2, "seuil", 3600, 7, "oui", "tenue", "", ""),
        (3, "vitesse", 3200, 8, "oui", "limite", "", ""),
        (4, "specifique", 4000, 8, "oui", "limite", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 7, 13), "allegee", [
        (0, "recuperation", 2200, 4, "oui", "tenue", "", "Deload"),
        (1, "technique", 2000, 4, "oui", "tenue", "", ""),
        (2, "endurance", 2400, 5, "oui", "tenue", "", ""),
        (4, "recuperation", 2200, 4, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 7, 20), "normale", [
        (0, "endurance", 3300, 6, "oui", "tenue", "", ""),
        (1, "technique", 2900, 5, "oui", "tenue", "", ""),
        (2, "seuil", 3500, 7, "oui", "tenue", "", ""),
        (3, "vitesse", 3100, 7, "oui", "tenue", "", ""),
        (4, "specifique", 3700, 7, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 7, 27), "normale", [
        (0, "endurance", 3500, 6, "oui", "tenue", "", ""),
        (1, "technique", 3000, 5, "oui", "tenue", "", ""),
        (2, "seuil", 3700, 7, "oui", "tenue", "", ""),
        (3, "vitesse", 3300, 8, "oui", "limite", "", ""),
        (4, "specifique", 4100, 8, "oui", "limite", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 8, 3), "normale", [
        (0, "endurance", 3600, 7, "oui", "tenue", "", ""),
        (1, "technique", 3200, 6, "oui", "tenue", "", ""),
        (2, "seuil", 3900, 8, "oui", "limite", "", ""),
        (3, "vitesse", 3400, 8, "oui", "casse", "", "Arrêt 200 m"),
        (4, "specifique", 4200, 8, "oui", "limite", "", "Volume +10%"),
    ])
    examples += _week_sessions("N002", date(2026, 8, 10), "allegee", [
        (0, "recuperation", 2400, 4, "oui", "tenue", "", "Deload"),
        (1, "technique", 2200, 4, "oui", "tenue", "", ""),
        (2, "endurance", 2600, 5, "oui", "tenue", "", ""),
        (4, "recuperation", 2400, 4, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N002", date(2026, 8, 17), "normale", [
        (0, "endurance", 3400, 6, "oui", "tenue", "", ""),
        (1, "technique", 3000, 5, "oui", "tenue", "", ""),
        (2, "seuil", 3600, 7, "oui", "tenue", "", "Allure 400"),
        (3, "vitesse", 3200, 7, "oui", "tenue", "", ""),
        (4, "specifique", 3800, 7, "oui", "tenue", "", ""),
    ])
    # Inès N003 — 2–3×/sem, T100 le 2 aout, T400 21 jun → B dû
    examples += _week_sessions("N003", date(2026, 6, 29), "normale", [
        (1, "technique", 1800, 5, "oui", "tenue", "", ""),
        (4, "endurance", 2200, 6, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N003", date(2026, 7, 6), "normale", [
        (1, "technique", 1900, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2300, 6, "oui", "tenue", "", ""),
        (5, "eau_libre", 2000, 6, "oui", "tenue", "", "Sighting"),
    ])
    examples += _week_sessions("N003", date(2026, 7, 13), "normale", [
        (1, "endurance", 2100, 6, "oui", "tenue", "", ""),
        (4, "seuil", 2400, 7, "oui", "limite", "", ""),
    ])
    examples += _week_sessions("N003", date(2026, 7, 20), "allegee", [
        (1, "recuperation", 1500, 4, "oui", "tenue", "", "Deload"),
        (4, "technique", 1600, 4, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N003", date(2026, 7, 27), "normale", [
        (1, "technique", 2000, 5, "oui", "tenue", "", ""),
        (5, "test", 1800, 7, "oui", "tenue", "T100", "T100 2 aout"),
    ])
    examples += _week_sessions("N003", date(2026, 8, 3), "normale", [
        (1, "endurance", 2200, 6, "oui", "tenue", "", ""),
        (3, "eau_libre", 2300, 6, "oui", "tenue", "", ""),
        (5, "specifique", 2100, 6, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N003", date(2026, 8, 10), "normale", [
        (1, "technique", 2100, 5, "oui", "tenue", "", ""),
        (3, "endurance", 2400, 6, "oui", "tenue", "", ""),
        (5, "eau_libre", 2200, 6, "oui", "tenue", "", ""),
    ])
    examples += _week_sessions("N003", date(2026, 8, 17), "normale", [
        (0, "eau_libre", 2200, 6, "oui", "tenue", "", "Sighting 25m"),
        (3, "endurance", 2300, 6, "oui", "tenue", "", ""),
        (4, "endurance", 2000, 6, "annulee", "casse", "", "Annulé orage"),
    ])

    last = 200
    input_cols = list(range(1, 11))
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            cell = input_cell(ws, r, c, val)
            if c in (10,):
                cell.alignment = left
        ws.cell(r, 2).number_format = "YYYY-MM-DD"
        formula_cell(ws, r, 11, f'=IF(B{r}="","",B{r}-WEEKDAY(B{r},3))', "YYYY-MM-DD")
        formula_cell(ws, r, 12, f'=IF(K{r}="","",YEAR(K{r})&"-W"&TEXT(ISOWEEKNUM(K{r}),"00"))')
        formula_cell(ws, r, 13, f'=IF(A{r}="","",A{r}&"|"&TEXT(K{r},"YYYY-MM-DD"))')
        ws.row_dimensions[r].height = 18
    start_empty = len(examples) + 2
    for r in range(start_empty, last + 1):
        for c in input_cols:
            input_cell(ws, r, c)
        ws.cell(r, 2).number_format = "YYYY-MM-DD"
        formula_cell(ws, r, 11, f'=IF(B{r}="","",B{r}-WEEKDAY(B{r},3))', "YYYY-MM-DD")
        formula_cell(ws, r, 12, f'=IF(K{r}="","",YEAR(K{r})&"-W"&TEXT(ISOWEEKNUM(K{r}),"00"))')
        formula_cell(ws, r, 13, f'=IF(A{r}="","",A{r}&"|"&TEXT(K{r},"YYYY-MM-DD"))')

    add_table(ws, "tblSeances", f"A1:M{last}")
    set_widths(ws, {
        "A": 12, "B": 12, "C": 14, "D": 11, "E": 8, "F": 10, "G": 14,
        "H": 10, "I": 12, "J": 28, "K": 12, "L": 12, "M": 22,
    })
    ws.freeze_panes = "C2"
    add_list_dv(ws, lists["theme_seance"], "C2:C200", "Thème")
    add_list_dv(ws, lists["faite_seance"], "F2:F200", "Séance faite ?")
    add_list_dv(ws, lists["type_semaine"], "G2:G200", "Type de semaine")
    add_list_dv(ws, lists["qualite"], "H2:H200", "Qualité")
    rpe = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    rpe.add("E2:E200")
    ws.add_data_validation(rpe)


def build_charge(wb: Workbook) -> None:
    ws = wb.create_sheet("CHARGE")
    ws.sheet_properties.tabColor = "3D8FFF"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.4)

    ws.merge_cells("A1:N1")
    ws["A1"] = (
        "CHARGE — 1 ligne = 1 nageur × 1 semaine. Ne rien saisir. "
        "Deload = 3 charges + 1 allégée (5–6 sem si 1–2 séances/sem). "
        "Retest A 4–5 sem, B 6–8 sem, jamais pendant deload ni < 3 sem avant l’objectif."
    )
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for col in range(1, 15):
        ws.cell(1, col).fill = fill_navy

    headers = [
        "id_nageur", "lundi", "nb_seances", "volume_m", "rpe_moy", "delta_vol",
        "type", "sem_deload", "sem_T100", "sem_T400", "sem_obj",
        "alerte_deload", "alerte_test", "consigne",
    ]
    header_row(ws, 2, headers, fill_blue)
    ws.row_dimensions[2].height = 32
    ws["F2"].comment = tip("vs semaine précédente. > 10 % = trop vite.")
    ws["H2"].comment = tip("Semaines depuis la dernière semaine allégée.")
    ws["L2"].comment = tip("due / en_cours / volume_+10% / rpe_2sem / trop_casse / affutage")
    ws["M2"].comment = tip("tester_A / tester_B / pas_maintenant")

    # A3:B# spill unique pairs, newest first
    ws["A3"] = (
        '=IFERROR(SORT(UNIQUE(FILTER(HSTACK(tblSeances[id_nageur],tblSeances[lundi]),'
        '(tblSeances[id_nageur]<>"")*(tblSeances[lundi]<>""))),2,-1),"")'
    )
    ws["A3"].fill = fill_formula
    ws["A3"].font = font_small
    ws["A3"].border = thin

    n_rows = 80
    for i in range(n_rows):
        r = 3 + i
        alt = i % 2 == 1

        def fc(c, val, fmt=None, wrap=False):
            if r == 3 and c in (1, 2):
                return
            formula_cell(ws, r, c, val, fmt, wrap)
            if alt:
                ws.cell(r, c).fill = fill_alt

        if r > 3:
            ws.cell(r, 1).border = thin
            ws.cell(r, 2).border = thin
            ws.cell(r, 2).number_format = "YYYY-MM-DD"

        fc(3, f'=IF($A{r}="","",COUNTIFS(tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r},tblSeances[faite],"oui"))')
        fc(4, f'=IF($A{r}="","",SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r},tblSeances[faite],"oui"))')
        fc(5, f'=IF($A{r}="","",IFERROR(AVERAGEIFS(tblSeances[rpe],tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r},tblSeances[faite],"oui"),""))', "0.0")
        fc(15, f'=IF($A{r}="","",SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r}-7,tblSeances[faite],"oui"))')  # O vol_prev
        fc(16, f'=IF($A{r}="","",IFERROR(AVERAGEIFS(tblSeances[rpe],tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r}-7,tblSeances[faite],"oui"),""))', "0.0")  # P rpe_prev
        fc(17, f'=IF($A{r}="","",COUNTIFS(tblSeances[id_nageur],$A{r},tblSeances[lundi],$B{r},tblSeances[qualite],"casse"))')  # Q
        fc(18, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[seances_sem],MATCH($A{r},tblNageurs[id],0)),3))')  # R freq
        fc(6, f'=IF(OR($A{r}="",O{r}="",O{r}=0),"", (D{r}-O{r})/O{r})', "0.0%")
        fc(7, (
            f'=IF($A{r}="","",IFERROR(INDEX(FILTER(tblSeances[type_semaine],'
            f'(tblSeances[id_nageur]=$A{r})*(tblSeances[lundi]=$B{r})*(tblSeances[type_semaine]<>"")),1),'
            f'IF(AND(O{r}>0,D{r}<=0.7*O{r}),"allegee","normale")))'
        ))
        fc(8, (
            f'=IF($A{r}="","",IFERROR(LET(d,MAXIFS(tblSeances[lundi],tblSeances[id_nageur],$A{r},'
            f'tblSeances[type_semaine],"allegee"),IF(d=0,"",($B{r}-d)/7)),""))'
        ), "0.0")
        fc(9, (
            f'=IF($A{r}="","",IFERROR(LET(d,MAXIFS(tblTests[date],tblTests[id_nageur],$A{r},'
            f'tblTests[protocole],"T100",tblTests[valide],"oui"),IF(d=0,"",($B{r}-d)/7)),""))'
        ), "0.0")
        fc(10, (
            f'=IF($A{r}="","",IFERROR(LET(d,MAXIFS(tblTests[date],tblTests[id_nageur],$A{r},'
            f'tblTests[protocole],"T400",tblTests[valide],"oui"),IF(d=0,"",($B{r}-d)/7)),""))'
        ), "0.0")
        fc(11, (
            f'=IF($A{r}="","",IFERROR((XLOOKUP($A{r},tblObjectifs[cle_dashboard],'
            f'tblObjectifs[date_objectif])-$B{r})/7,""))'
        ), "0.0")
        fc(12, (
            f'=IF($A{r}="","",IFS('
            f'C{r}<=1,"",'
            f'AND(K{r}<>"",K{r}<=3),"affutage",'
            f'G{r}="allegee","en_cours",'
            f'AND(H{r}<>"",H{r}>=IF(R{r}<=2,5,3),C{r}>=IF(R{r}<=2,2,3)),"due",'
            f'AND(F{r}<>"",F{r}>0.1),"volume_+10%",'
            f'AND(E{r}<>"",E{r}>=8,P{r}<>"",P{r}>=8),"rpe_2sem",'
            f'Q{r}>=2,"trop_casse",'
            f'TRUE,""))'
        ))
        fc(13, (
            f'=IF($A{r}="","",IFS('
            f'AND(K{r}<>"",K{r}<=3),"pas_maintenant",'
            f'G{r}="allegee","attendre_fin_deload",'
            f'AND(E{r}<>"",E{r}>=8),"pas_maintenant",'
            f'AND(I{r}<>"",I{r}>=5,J{r}<>"",J{r}>=7),"tester_A_et_B",'
            f'AND(I{r}<>"",I{r}>=5),"tester_A",'
            f'AND(J{r}<>"",J{r}>=7),"tester_B",'
            f'TRUE,""))'
        ))
        fc(14, (
            f'=IF($A{r}="","",TEXTJOIN(" · ",TRUE,'
            f'IF(L{r}="due","Deload cette semaine (−30% vol, RPE≤6, pas de max)",""),'
            f'IF(L{r}="en_cours","Deload en cours — garder le leger",""),'
            f'IF(L{r}="volume_+10%","Volume +10% vs S-1 — alleger S+1",""),'
            f'IF(L{r}="rpe_2sem","RPE eleve 2 sem — alleger maintenant",""),'
            f'IF(L{r}="trop_casse","Trop de seances cassees — recup cette semaine",""),'
            f'IF(L{r}="affutage","Affutage : plus de max, allure de course",""),'
            f'IF(M{r}="tester_A","Retest A (T50+T100) apres une seance facile",""),'
            f'IF(M{r}="tester_B","Retest B (T200+T400) a frais",""),'
            f'IF(M{r}="tester_A_et_B","Retest A cette sem, B la suivante",""),'
            f'IF(M{r}="attendre_fin_deload","Retest apres le deload, pas pendant",""),'
            f'IF(M{r}="pas_maintenant","Pas de test a fond",""),'
            f'IF(AND(L{r}="",M{r}=""),"Charge OK — semaine normale","")))'
        ), wrap=True)
        ws.cell(r, 14).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 28

    ws.cell(3, 2).number_format = "YYYY-MM-DD"
    for col in range(15, 19):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    set_widths(ws, {
        "A": 12, "B": 12, "C": 12, "D": 12, "E": 10, "F": 11, "G": 12,
        "H": 12, "I": 11, "J": 11, "K": 10, "L": 14, "M": 18, "N": 52,
    })
    ws.freeze_panes = "C3"
    ws.conditional_formatting.add("L3:L82", FormulaRule(formula=['L3="due"'], fill=fill_red))
    ws.conditional_formatting.add("L3:L82", FormulaRule(formula=['L3="en_cours"'], fill=fill_amber))
    ws.conditional_formatting.add("L3:L82", FormulaRule(formula=['L3="volume_+10%"'], fill=fill_amber))
    ws.conditional_formatting.add("L3:L82", FormulaRule(formula=['L3="affutage"'], fill=fill_green))
    ws.conditional_formatting.add("M3:M82", FormulaRule(formula=['LEFT(M3,6)="tester"'], fill=fill_amber))
    ws.conditional_formatting.add("F3:F82", CellIsRule(operator="greaterThan", formula=["0.1"], fill=fill_amber))
    ws.conditional_formatting.add("C3:C82", CellIsRule(operator="greaterThan", formula=["5"], fill=fill_red))

    ws.merge_cells("A84:N84")
    ws["A84"] = (
        "Saisie uniquement dans SEANCES. Ici tout est calculé. "
        "nb_seances = séances faites (oui). Max 5 / semaine. "
        "Exemples N001–N003 : 8 semaines d’historique pour voir deload et retest."
    )
    ws["A84"].font = font_muted


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("DASHBOARD", 1)
    ws.sheet_properties.tabColor = NAVY
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.55, bottom=0.4)

    ws.merge_cells("A1:AC1")
    ws["A1"] = "TABLEAU DE BORD — tests · charge · deload · retest · ne rien saisir ici"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    for col in range(1, 30):
        ws.cell(1, col).fill = fill_navy

    headers = [
        "id", "prenom", "nom", "niveau", "statut", "nage_suivi", "discipline",
        "nage_obj", "dist_m", "date_obj", "lieu", "cible", "semaines",
        "T50", "T100", "T200", "T400", "delta_T100", "drop_100/50",
        "drop_400/100", "CSS_/100", "T100_sec", "diagnostic",
        "seances_S", "vol_S", "vol_S/S-1/S-2", "deload", "retest", "consigne_charge",
    ]
    header_row(ws, 2, headers, fill_blue)
    ws.row_dimensions[2].height = 34
    ws["R2"].comment = tip("Négatif = plus rapide qu’au test précédent.")
    ws["S2"].comment = tip("T100 vs 2×T50. > 10 % : il casse sur 100.")
    ws["T2"].comment = tip("T400 vs 4×T100. > 12 % : manque d’aérobie.")
    ws["U2"].comment = tip("CSS = (T400 − T200) / 2. Allure seuil / 100 m.")
    ws["V2"].comment = tip("Dernier T100 valide en secondes — pace100 MySWYM.")
    ws["X2"].comment = tip("Séances faites cette semaine / cible nageur.")
    ws["Z2"].comment = tip("Volumes S / S-1 / S-2 (mètres).")
    ws["AA2"].comment = tip("Voir CHARGE. due = alléger cette semaine.")
    ws["AB2"].comment = tip("tester_A = T50+T100 · tester_B = T200+T400.")

    n_rows = 22
    for i in range(n_rows):
        r = 3 + i
        idx = i + 1
        nage = f"$F{r}"
        alt = i % 2 == 1

        def paint(c, val, fmt=None, wrap=False):
            formula_cell(ws, r, c, val, fmt, wrap)
            if alt:
                ws.cell(r, c).fill = fill_alt

        paint(1, f'=IFERROR(IF(INDEX(tblNageurs[id],{idx})="","",INDEX(tblNageurs[id],{idx})),"")')
        paint(2, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[prenom],{idx}),""))')
        paint(3, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[nom],{idx}),""))')
        paint(4, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[niveau],{idx}),""))')
        paint(5, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[statut],{idx}),""))')
        paint(7, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[discipline],""),""))')
        paint(8, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[nage],""),""))')
        paint(6, f'=IF($A{r}="","",IF(OR($H{r}="dos",$H{r}="brasse",$H{r}="papillon"),$H{r},"crawl"))')
        paint(9, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[distance_m],""),""))')
        paint(10, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[date_objectif],""),""))', "YYYY-MM-DD")
        paint(11, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[lieu],""),""))')
        ws.cell(r, 11).alignment = left
        paint(12, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[temps_cible],""),""))')
        paint(13, f'=IF(OR($A{r}="",$J{r}=""),"",($J{r}-TODAY())/7)', "0.0")

        formula_cell(ws, r, 53, f"={last_sec(r, 'T50', nage)}", "0.00")
        formula_cell(ws, r, 54, f"={last_sec(r, 'T100', nage)}", "0.00")
        formula_cell(ws, r, 55, f"={last_sec(r, 'T200', nage)}", "0.00")
        formula_cell(ws, r, 56, f"={last_sec(r, 'T400', nage)}", "0.00")
        formula_cell(ws, r, 57, f"={prev_sec(r, 'T100', nage)}", "0.00")
        formula_cell(ws, r, 58, f"={last_date(r, 'T100', nage)}", "YYYY-MM-DD")

        paint(14, f"={fmt_temps(f'BA{r}')}")
        paint(15, f"={fmt_temps(f'BB{r}')}")
        paint(16, f"={fmt_temps(f'BC{r}')}")
        paint(17, f"={fmt_temps(f'BD{r}')}")
        paint(18, f'=IF(OR(BB{r}="",BE{r}=""),"",(BB{r}-BE{r})/BE{r})', "0.0%")
        paint(19, f'=IF(OR(BA{r}="",BB{r}=""),"",BB{r}/(2*BA{r})-1)', "0.0%")
        paint(20, f'=IF(OR(BB{r}="",BD{r}=""),"",BD{r}/(4*BB{r})-1)', "0.0%")
        paint(21, f'=IF(OR(BC{r}="",BD{r}=""),"",INT(((BD{r}-BC{r})/2)/60)&":"&TEXT(MOD((BD{r}-BC{r})/2,60),"00.00"))')
        paint(22, f'=IF(BB{r}="","",ROUND(BB{r},1))', "0.0")
        diag = (
            f'=IF($A{r}="","",IFS('
            f'BB{r}="","Pas de T100 valide — Batterie A (T50 + T100).",'
            f'AND(M{r}<>"",M{r}<=3),"Affutage (< 3 sem) — plus de test a fond. Allure de course.",'
            f'AND(S{r}<>"",S{r}>0.1),"Vitesse OK, il casse sur 100 (lactique / pacing / technique).",'
            f'AND(T{r}<>"",T{r}>0.12),"Manque aerobie / seuil — Batterie B (T200+T400) et series CSS.",'
            f'AND(T{r}<>"",T{r}<0.06,S{r}<>"",S{r}<0.08),"Profil homogene — specificite d\'allure vers la cible.",'
            f'BA{r}="","Ajouter un T50 a la prochaine Batterie A.",'
            f'TRUE,"Profil coherent. Voir consigne_charge pour deload / retest."))'
        )
        paint(23, diag, wrap=True)
        ws.cell(r, 23).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        lundi = f"(TODAY()-WEEKDAY(TODAY(),3))"
        paint(24, (
            f'=IF($A{r}="","",COUNTIFS(tblSeances[id_nageur],$A{r},tblSeances[lundi],{lundi},'
            f'tblSeances[faite],"oui")&"/"&IFERROR(INDEX(tblNageurs[seances_sem],MATCH($A{r},tblNageurs[id],0)),""))'
        ))
        paint(25, (
            f'=IF($A{r}="","",SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},'
            f'tblSeances[lundi],{lundi},tblSeances[faite],"oui"))'
        ), "#,##0")
        paint(26, (
            f'=IF($A{r}="","",TEXT(SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},'
            f'tblSeances[lundi],{lundi},tblSeances[faite],"oui"),"0")&" / "&'
            f'TEXT(SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},'
            f'tblSeances[lundi],{lundi}-7,tblSeances[faite],"oui"),"0")&" / "&'
            f'TEXT(SUMIFS(tblSeances[volume_m],tblSeances[id_nageur],$A{r},'
            f'tblSeances[lundi],{lundi}-14,tblSeances[faite],"oui"),"0"))'
        ))
        paint(27, (
            f'=IF($A{r}="","",IFERROR(INDEX(FILTER(CHARGE!L$3:L$82,'
            f'(CHARGE!A$3:A$82=$A{r})*(CHARGE!B$3:B$82={lundi})),1),""))'
        ))
        paint(28, (
            f'=IF($A{r}="","",IFERROR(INDEX(FILTER(CHARGE!M$3:M$82,'
            f'(CHARGE!A$3:A$82=$A{r})*(CHARGE!B$3:B$82={lundi})),1),""))'
        ))
        paint(29, (
            f'=IF($A{r}="","",IFERROR(INDEX(FILTER(CHARGE!N$3:N$82,'
            f'(CHARGE!A$3:A$82=$A{r})*(CHARGE!B$3:B$82={lundi})),1),""))'
        ), wrap=True)
        ws.cell(r, 29).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 36

    for col in range(53, 59):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    set_widths(ws, {
        "A": 8, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 14, "H": 10,
        "I": 9, "J": 12, "K": 22, "L": 10, "M": 10, "N": 9, "O": 9, "P": 9,
        "Q": 9, "R": 11, "S": 11, "T": 12, "U": 10, "V": 10, "W": 36,
        "X": 11, "Y": 10, "Z": 18, "AA": 14, "AB": 16, "AC": 42,
    })
    ws.freeze_panes = "F3"
    ws.auto_filter.ref = f"A2:AC{2 + n_rows}"

    ws.conditional_formatting.add("M3:M24", CellIsRule(operator="lessThan", formula=["3"], fill=fill_red))
    ws.conditional_formatting.add("M3:M24", CellIsRule(operator="between", formula=["3", "8"], fill=fill_amber))
    ws.conditional_formatting.add("M3:M24", CellIsRule(operator="greaterThan", formula=["8"], fill=fill_green))
    ws.conditional_formatting.add("R3:R24", CellIsRule(operator="lessThan", formula=["0"], fill=fill_green))
    ws.conditional_formatting.add("R3:R24", CellIsRule(operator="greaterThan", formula=["0.015"], fill=fill_red))
    ws.conditional_formatting.add("S3:S24", CellIsRule(operator="greaterThan", formula=["0.1"], fill=fill_amber))
    ws.conditional_formatting.add("T3:T24", CellIsRule(operator="greaterThan", formula=["0.12"], fill=fill_amber))
    ws.conditional_formatting.add("E3:E24", FormulaRule(formula=['E3="actif"'], fill=fill_green))
    ws.conditional_formatting.add("E3:E24", FormulaRule(formula=['E3="pause"'], fill=fill_grey))
    ws.conditional_formatting.add("AA3:AA24", FormulaRule(formula=['AA3="due"'], fill=fill_red))
    ws.conditional_formatting.add("AA3:AA24", FormulaRule(formula=['AA3="en_cours"'], fill=fill_amber))
    ws.conditional_formatting.add("AB3:AB24", FormulaRule(formula=['LEFT(AB3,6)="tester"'], fill=fill_amber))

    ws.merge_cells("A26:AC26")
    ws["A26"] = (
        "Δ T100 vert = progrès. Drop 100/50 > 10 % ou 400/100 > 12 % = levier. "
        "Deload due = alléger cette semaine. tester_A / tester_B = retest (jamais pendant deload). "
        "Détail semaine par semaine dans CHARGE. Exemples N001–N003 à remplacer."
    )
    ws["A26"].font = font_muted


def build_export(wb: Workbook) -> None:
    ws = wb.create_sheet("EXPORT")
    ws.sheet_properties.tabColor = "5B6B7C"
    headers = [
        "id", "prenom", "nom", "niveau", "pool", "sessionsPerWeek", "objectifV1",
        "raceDistance", "raceStroke", "raceTargetTimeSec", "competitionDate",
        "lieu", "pace100", "injuryStatus", "injuryNote",
    ]
    header_row(ws, 1, headers)
    ws["M1"].comment = tip("Dernier T100 valide (secondes) = pace100 dans l’app.")
    n_rows = 22
    for i in range(n_rows):
        r = 2 + i
        idx = i + 1
        formula_cell(ws, r, 1, f'=IFERROR(IF(INDEX(tblNageurs[id],{idx})="","",INDEX(tblNageurs[id],{idx})),"")')
        formula_cell(ws, r, 2, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[prenom],{idx}),""))')
        formula_cell(ws, r, 3, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[nom],{idx}),""))')
        formula_cell(ws, r, 4, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[niveau],{idx}),""))')
        formula_cell(ws, r, 5, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[bassin],{idx}),""))')
        formula_cell(ws, r, 6, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[seances_sem],{idx}),""))')
        formula_cell(ws, r, 7, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[discipline],""),""))')
        formula_cell(ws, r, 8, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[distance_m],""),""))')
        formula_cell(ws, r, 9, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[nage],""),""))')
        formula_cell(ws, r, 10, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[temps_cible_sec],""),""))', "0.00")
        formula_cell(ws, r, 11, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[date_objectif],""),""))', "YYYY-MM-DD")
        formula_cell(ws, r, 12, f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},tblObjectifs[cle_dashboard],tblObjectifs[lieu],""),""))')
        formula_cell(ws, r, 13, f'=IF($A{r}="","",IFERROR(INDEX(DASHBOARD!V:V,MATCH($A{r},DASHBOARD!A:A,0)),""))', "0.0")
        formula_cell(ws, r, 14, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[blessure],{idx}),""))')
        formula_cell(ws, r, 15, f'=IF($A{r}="","",IFERROR(INDEX(tblNageurs[note_contrainte],{idx}),""))')
        ws.row_dimensions[r].height = 20
    set_widths(ws, {
        "A": 8, "B": 12, "C": 14, "D": 14, "E": 8, "F": 16, "G": 16,
        "H": 14, "I": 12, "J": 18, "K": 16, "L": 28, "M": 12, "N": 14, "O": 22,
    })
    ws.freeze_panes = "A2"
    ws.merge_cells("A25:O25")
    ws["A25"] = "Fichier → Enregistrer sous → CSV UTF-8 si import plus tard. Ne saisis rien ici."
    ws["A25"].font = font_muted


def build_evolution(wb: Workbook) -> None:
    ws = wb.create_sheet("EVOLUTION")
    ws.sheet_properties.tabColor = "3D8FFF"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Historique T100 valides (se remplit tout seul) + courbe d’exemple Léa N001"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = left
    for col in range(1, 7):
        ws.cell(1, col).fill = fill_navy
    header_row(ws, 2, ["id_nageur", "prenom", "date", "nage", "temps_sec", "temps"])
    formula_cell(ws, 3, 1, '=IFERROR(FILTER(tblTests[id_nageur],(tblTests[protocole]="T100")*(tblTests[valide]="oui")),"")')
    formula_cell(ws, 3, 2, '=IF(A3="","",IFERROR(XLOOKUP(A3,tblNageurs[id],tblNageurs[prenom]),""))')
    formula_cell(ws, 3, 3, '=IFERROR(FILTER(tblTests[date],(tblTests[protocole]="T100")*(tblTests[valide]="oui")),"")', "YYYY-MM-DD")
    formula_cell(ws, 3, 4, '=IFERROR(FILTER(tblTests[nage],(tblTests[protocole]="T100")*(tblTests[valide]="oui")),"")')
    formula_cell(ws, 3, 5, '=IFERROR(FILTER(tblTests[temps_sec],(tblTests[protocole]="T100")*(tblTests[valide]="oui")),"")', "0.00")
    formula_cell(ws, 3, 6, f"={fmt_temps('E3')}")

    ws.merge_cells("H1:I1")
    ws["H1"] = "N001 T100 crawl — source graphique"
    ws["H1"].font = font_h
    ws["H1"].fill = fill_blue
    ws["I1"].fill = fill_blue
    ws["H2"] = "date"
    ws["I2"] = "T100 (s)"
    ws["H2"].font = font_h
    ws["H2"].fill = fill_navy
    ws["I2"].font = font_h
    ws["I2"].fill = fill_navy
    pts = [
        (date(2026, 3, 15), 76.2),
        (date(2026, 5, 10), 74.8),
        (date(2026, 7, 20), 73.4),
    ]
    for i, (d, sec) in enumerate(pts, 3):
        input_cell(ws, i, 8, d).number_format = "YYYY-MM-DD"
        input_cell(ws, i, 9, sec).number_format = "0.00"

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Lea Martin — T100 crawl (exemple)"
    chart.y_axis.title = "secondes"
    data = Reference(ws, min_col=9, min_row=2, max_row=5)
    cats = Reference(ws, min_col=8, min_row=3, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.y_axis.scaling.min = 70
    chart.width = 14
    chart.height = 8
    ws.add_chart(chart, "H7")

    set_widths(ws, {
        "A": 12, "B": 12, "C": 14, "D": 12, "E": 12, "F": 12, "H": 14, "I": 12,
    })
    ws.freeze_panes = "A3"
    ws.merge_cells("A20:F22")
    ws["A20"] = (
        "Le tableau A:F se remplit avec FILTER (Excel 365) dès que tu ajoutes des T100 valides. "
        "Le graphique d’exemple est calé sur Léa (N001, points H3:I5). "
        "Pour un autre nageur, change les dates/temps dans H:I d’après l’onglet TESTS."
    )
    ws["A20"].font = font_muted
    ws["A20"].alignment = left_top


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)
    wb.properties.title = "Suivi nageurs — coach de ligne"
    wb.properties.creator = "MySWYM"
    wb.properties.description = "Tests, séances 1–5/sem, charge, deload, retest, export app."

    lists = build_listes(wb)
    build_guide(wb)
    build_nageurs(wb, lists)
    build_objectifs(wb, lists)
    build_tests(wb, lists)
    build_seances(wb, lists)
    build_charge(wb)
    build_dashboard(wb)
    build_export(wb)
    build_evolution(wb)

    order = ["GUIDE", "DASHBOARD", "NAGEURS", "OBJECTIFS", "TESTS", "SEANCES", "CHARGE", "EVOLUTION", "EXPORT", "LISTES"]
    for i, name in enumerate(order):
        current = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - current)

    wb.save(OUT)
    print("Wrote", OUT)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
